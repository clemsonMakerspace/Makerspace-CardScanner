"""
Excel utilities with corruption prevention measures
Implements file locking, backups, and safe write operations
"""
import os
import shutil
import time
import threading
from datetime import datetime
from openpyxl import load_workbook
from contextlib import contextmanager

# Global lock for Excel file access
_excel_lock = threading.Lock()

# Backup configuration
BACKUP_DIR = "backups"  # Use existing backup directory
MAX_BACKUPS = 50  # Keep last 50 backups
BACKUP_INTERVAL = 36000  # Create backup every 10 hours (in seconds)
_last_backup_time = 0
BACKUP_TIMESTAMP_FILE = os.path.join(BACKUP_DIR, ".last_backup_time")  # Persistent timestamp file

def ensure_backup_dir():
    """Create backup directory if it doesn't exist"""
    if not os.path.exists(BACKUP_DIR):
        os.makedirs(BACKUP_DIR)
        print(f"Created backup directory: {BACKUP_DIR}")

def create_backup(file_path, location="Watt"):
    """
    Create a timestamped backup of the Excel file
    Only creates backup if enough time has passed since last backup
    Uses persistent file to track last backup time across process restarts
    Matches existing backup naming convention: hardware_users_{Location}_{date}.xlsx
    """
    global _last_backup_time
    
    ensure_backup_dir()
    
    current_time = time.time()
    
    # Load last backup time from file if it exists
    if _last_backup_time == 0 and os.path.exists(BACKUP_TIMESTAMP_FILE):
        try:
            with open(BACKUP_TIMESTAMP_FILE, 'r') as f:
                _last_backup_time = float(f.read().strip())
        except:
            _last_backup_time = 0
    
    # Check if enough time has passed since last backup
    time_since_backup = current_time - _last_backup_time
    if time_since_backup < BACKUP_INTERVAL:
        # Skip backup - not enough time has passed
        hours_remaining = (BACKUP_INTERVAL - time_since_backup) / 3600
        print(f"[DEBUG] Backup skipped (next backup in {hours_remaining:.1f} hours)")  # TEMP DEBUG
        return None
    
    try:
        # Match existing format: MM_DD_YYYY___HH_MM_SS
        timestamp = datetime.now().strftime("%m_%d_%Y___%H_%M_%S")
        backup_filename = f"hardware_users_{location}_{timestamp}.xlsx"
        backup_path = os.path.join(BACKUP_DIR, backup_filename)
        
        # Copy the file
        shutil.copy2(file_path, backup_path)
        _last_backup_time = current_time
        
        # Save timestamp to file for persistence across process restarts
        with open(BACKUP_TIMESTAMP_FILE, 'w') as f:
            f.write(str(current_time))
        
        print(f"✓ Backup created: {backup_filename}")
        
        # Clean up old backups
        cleanup_old_backups()
        
        return backup_path
    except Exception as e:
        print(f"Warning: Failed to create backup: {e}")
        return None

def cleanup_old_backups():
    """Remove old backups, keeping only MAX_BACKUPS most recent"""
    try:
        backup_files = [
            os.path.join(BACKUP_DIR, f) 
            for f in os.listdir(BACKUP_DIR) 
            if f.startswith("hardware_users_") and f.endswith(".xlsx")
        ]
        
        # Sort by modification time (oldest first)
        backup_files.sort(key=os.path.getmtime)
        
        # Remove oldest backups if we have too many
        while len(backup_files) > MAX_BACKUPS:
            oldest = backup_files.pop(0)
            os.remove(oldest)
            print(f"Removed old backup: {os.path.basename(oldest)}")
    except Exception as e:
        print(f"Warning: Failed to cleanup old backups: {e}")

@contextmanager
def safe_excel_write(file_path, max_retries=3, retry_delay=0.5):
    """
    Context manager for safe Excel file writing with locking and retries
    
    Usage:
        with safe_excel_write(file_path) as wb:
            sheet = wb['Users']
            sheet['A1'] = 'value'
            # File is automatically saved when exiting context
    """
    workbook = None
    retries = 0
    
    while retries < max_retries:
        try:
            # Acquire lock
            _excel_lock.acquire()
            
            # Create backup periodically
            create_backup(file_path)
            
            # Load workbook
            workbook = load_workbook(filename=file_path)
            
            # Yield workbook to caller
            yield workbook
            
            # Save workbook with temp file + rename for atomic write
            temp_path = file_path + '.tmp'
            workbook.save(temp_path)
            
            # Atomic rename (reduces corruption risk)
            if os.path.exists(file_path):
                os.replace(temp_path, file_path)
            else:
                os.rename(temp_path, file_path)
            
            # Success - break out of retry loop
            break
            
        except PermissionError as e:
            retries += 1
            if retries >= max_retries:
                print(f"ERROR: Could not access Excel file after {max_retries} attempts: {e}")
                raise
            print(f"File locked, retrying ({retries}/{max_retries})...")
            time.sleep(retry_delay)
            
        except Exception as e:
            print(f"ERROR during Excel write: {e}")
            # Clean up temp file if it exists
            temp_path = file_path + '.tmp'
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except:
                    pass
            raise
            
        finally:
            # Always release lock
            if _excel_lock.locked():
                _excel_lock.release()
            
            # Close workbook
            if workbook:
                try:
                    workbook.close()
                except:
                    pass

@contextmanager
def safe_excel_read(file_path, max_retries=3, retry_delay=0.5):
    """
    Context manager for safe Excel file reading with retries
    
    Usage:
        with safe_excel_read(file_path) as wb:
            sheet = wb['Users']
            value = sheet['A1'].value
    """
    workbook = None
    retries = 0
    
    while retries < max_retries:
        try:
            # Read-only access doesn't need lock (openpyxl reads into memory)
            workbook = load_workbook(filename=file_path, read_only=False, data_only=True)
            yield workbook
            break
            
        except PermissionError as e:
            retries += 1
            if retries >= max_retries:
                print(f"ERROR: Could not read Excel file after {max_retries} attempts: {e}")
                raise
            print(f"File locked, retrying read ({retries}/{max_retries})...")
            time.sleep(retry_delay)
            
        except Exception as e:
            print(f"ERROR during Excel read: {e}")
            raise
            
        finally:
            if workbook:
                try:
                    workbook.close()
                except:
                    pass

def verify_excel_integrity(file_path):
    """
    Verify that Excel file can be opened and is not corrupted
    Returns True if file is OK, False if corrupted
    """
    try:
        with safe_excel_read(file_path) as wb:
            # Try to access basic properties
            sheet_names = wb.sheetnames
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                _ = sheet.max_row
                _ = sheet.max_column
        return True
    except Exception as e:
        print(f"Excel file appears corrupted: {e}")
        return False

def restore_from_backup(file_path, backup_path=None):
    """
    Restore Excel file from backup
    If backup_path is None, uses most recent backup
    """
    ensure_backup_dir()
    
    if backup_path is None:
        # Find most recent backup (matches existing format: hardware_users_*.xlsx)
        backup_files = [
            os.path.join(BACKUP_DIR, f) 
            for f in os.listdir(BACKUP_DIR) 
            if f.startswith("hardware_users_") and f.endswith(".xlsx")
        ]
        
        if not backup_files:
            print("ERROR: No backups found!")
            return False
        
        # Sort by modification time (newest first)
        backup_files.sort(key=os.path.getmtime, reverse=True)
        backup_path = backup_files[0]
    
    try:
        # Move corrupted file to a safe location
        corrupted_path = file_path + '.corrupted.' + datetime.now().strftime("%Y%m%d_%H%M%S")
        shutil.move(file_path, corrupted_path)
        print(f"Moved corrupted file to: {corrupted_path}")
        
        # Restore from backup
        shutil.copy2(backup_path, file_path)
        print(f"✓ Restored from backup: {os.path.basename(backup_path)}")
        
        return True
    except Exception as e:
        print(f"ERROR: Failed to restore from backup: {e}")
        return False
