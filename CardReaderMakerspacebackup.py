import sys
import customtkinter as ctk
from openpyxl import load_workbook
import tkinter as tk
from tkinter import simpledialog, Canvas, messagebox
import PIL
from PIL import Image, ImageTk 
from screeninfo import get_monitors
import pygetwindow as gw
from datetime import datetime
from __main__ import *
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
from pynput.keyboard import Key, Controller
import time
import requests


###    Version 1.2

#Path to excel sheet
file_path = "hardware_users.xlsx"
sheet_name = "Scans"
sheet2_name ="Users"
Location= "Watt"
keyboard = Controller()

def load_excel(): #this is a bit redundant but it works, can eventually use this in "add user to shee"
    # Load the workbook and sheet
    workbook = load_workbook(filename=file_path)
    sheet = workbook[sheet_name] 
    sheet2 = workbook[sheet2_name] 
    return workbook, sheet, sheet2

#def load_hardware_ids(sheet): #Not even using this rn
    hardware_dict = {}
    # Load the hardware IDs and usernames into a dictionary
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        hardware_dict[str(row[0])] = row[1]  # Hardware ID as the key, username as the value

def find_hardware_id(sheet2, hardware_id):
    # Loop through the rows starting from row 2 to skip headers (if any)
    for row in sheet2.iter_rows(min_row=2, max_row=sheet2.max_row, values_only=True):
        if str(row[1]) == str(hardware_id):  # Compare hardware_id in column B (index 1)
            return row[0]  # Return the username from column A (index 0)
    
    return None  # Return None if not found
def find_userdata(hardware_id,sheet2):
    # Loop through the rows starting from row 2 to skip headers
    for row in sheet2.iter_rows(min_row=2, max_row=sheet2.max_row, values_only=True):
        if str(row[1]) == str(hardware_id):  # Look for hardware_id in col. B
            first_name = row[3]  # Column D 
            last_name = row[4]   # Column E 
            major = row[5]       # Column F
            return first_name, last_name, major
    
    return None, None, None  #   Set to None if they are not found

def add_user_to_sheet(sheet_name,sheet2_name,hardware_id, username,first_name,last_name,major,workbook,userstatus):
    wb = workbook
    scans_sheet = wb[sheet_name]
    users_sheet = wb[sheet2_name]
    
    if userstatus == 1:
            # Search for matching hardware ID in the "Users" sheet or for an empty hardware ID cell
            for row in users_sheet.iter_rows(min_row=2, values_only=False):  # Skip header row
                cell_hardware_id = row[1].value  # Column B in "Users" for hardware_id

            # Cast both the hardware_id from input and the one from the sheet to str for comparison
                if str(cell_hardware_id) == str(hardware_id): #str is irrelavent but may help edge cases where numbers get input as strings?
                    match_found = True
                    print(f"User with hardware ID {hardware_id} already exists in 'Users' sheet.")
                    break  # Stop searching after finding the match

                # If the hardware ID cell is empty (i.e., new entry row), fill in this row
                if cell_hardware_id is None or cell_hardware_id == "":  # Check for an empty hardware ID
                    row[0].value = username  # Column A for username
                    row[1].value = int(hardware_id)  # Column B for hardware ID
                    row[3].value = first_name  # Column D for first name
                    row[4].value = last_name   # Column E for last name
                    row[5].value = major       # Column F for major
                    match_found = True
                    print(f"New user {first_name} {last_name} added to 'Users' sheet.")
                    break  # Stop searching after appending the new data

    # Add the scan to the 'Scans' sheet (this happens regardless of userstatus)
    now = datetime.now()
    #timestamp = now.timestamp() #use this for seconds only
    timestamp = now.strftime('%m/%d/%Y %H:%M:%S') # Format the time to display as "YYYY-MM-DD HH:MM"
    scans_sheet.append([int(hardware_id), username, timestamp])
    
    # Save the workbook after making changes
    wb.save(file_path)
    print(f"Scan Added, workbook saved.")
#def api_call(hardware_id,Location,timestamp):
def scrape_user(username):
    # Set up Selenium with headless Chrome
        chrome_options = Options()
        chrome_options.add_argument("--headless")  # Run Chrome in headless mode
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")

        # Set up the driver
        driver = webdriver.Chrome(options=chrome_options)

        # Create da URL using username, change this id the directory changes url
        url = f"https://my.clemson.edu/#/directory/person/{username}"

        # Load da page
        driver.get(url)

        # Wait for the full name element to appear
        try:
            WebDriverWait(driver, 1).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, '.personView .primaryInfo h2')) 
            )
            #the time should be upped if the scrape does not complete in time
        except TimeoutException:
            print(f"Timeout while waiting for the element on the page for {username}.")
            driver.quit()
            return username, None, None  # Default values if element not found
        except Exception as e:
            print(f"Error during page load: {e}")
            driver.quit()
            return username, None, None  # Default values if there's an error
        # Get da page
        page_source = driver.page_source

        # Parse the page with BeautifulSoup
        soup = BeautifulSoup(page_source, 'html.parser')

        #Find da Name
        full_name_element = soup.select_one('.personView .primaryInfo h2') 
        if full_name_element:
            full_name = full_name_element.get_text().strip()

            # Split the full name by spaces
            name_parts = full_name.split()

            # Store the first name and last name (ignore middle name if present)
            first_name = name_parts[0]  # The first part of the name
            last_name = name_parts[-1]  # The last part of the name, if they have like a number or sr. or something after their last name itll fuck it up but it doesn't matter because we don't really need the last name that badly anyways
        else:
            first_name, last_name = None, None
            
        #Find major (sometimes selects other stuff on page cuz directory only has some of the stuff sometimes)
        major_element = soup.select_one('.personView .primaryInfo .data p')
        if major_element:
            major = major_element.get_text().strip()
        else:
            major = None
        # Print the scraped information
        print(f"First Name: {first_name}")
        print(f"Last Name: {last_name}")
        print(f"Major: {major}")

        # Close the scraping driver/borderless window
        driver.quit()
        return first_name, last_name, major
'''def api_call(hardware_id,Location,username)
    timestamp = now.strftime('%Y-%d-%mT%H:%M:%S') #Format the time to display like "2024-01-09T15:49:00" for the AWS database dictonary format.
    data = {
        "user_id": {hardware_id},
        "location": {Location},
        "timestamp": "2024-01-09T15:49:00",
    }

    headers = {
        "x-api-key": "bY2BQ0boppPn3rnfSjsh68kgQKpBYMq4eP5uWLvd"
    }

    requests.post("https://ney0lua6fb.execute-api.us-east-1.amazonaws.com/development/visits", headers=headers, data=data)
    '''
def make_fullscreen_on_top(root):
    root.attributes('-fullscreen', True)
    root.attributes('-topmost', True)
#this can probably be put within another function but it works for now

def show_welcome_popup(root, username, first_name, userstatus):
    # Set the background image
    image = Image.open("backgroundLarge.png")
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    image = image.resize((screen_width, screen_height), PIL.Image.Resampling.LANCZOS)
    bg_image = ImageTk.PhotoImage(image)
    root.bg_image = bg_image  # Keep a reference to avoid garbage collection
    background_label = tk.Label(root, image=bg_image)   # Create a label for the background
    background_label.place(relwidth=1, relheight=1)  # Stretch to fit window (Idek if this works properly because it isn't doing it that well)
    
    # welcome back message
    if first_name is None:
        first_name = username
    if userstatus == 0: #Returning user
        message = (f"Welcome back, {first_name}!")
    else: # New User
        message = (f"Welcome to the Makerspace! Thank you!")
    message_label = tk.Label(root, text=message, font=("Helvetica", 40, "bold"), fg="white", bg="black")
    message_label.place(relx=0.5, rely=0.5, anchor="center")  # Center the message

    root.after(2500, root.quit)  # Close the window after 2.5 seconds

def close_on_escape(event): #this is maybe redundant because I do this induvidually within some functions
    print("Escape key pressed. Exiting program...") #esc to close
    sys.exit()  # Exit the program

def prompt_for_username():
    # Simple dialog to ask for a username
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    root.after(25000, root.quit)  # Close the window after 35 seconds
    root.attributes('-fullscreen', True)  # Make the window full screen
    #root.focus_force()
    
    def is_valid_username(entered_username):
        """Validate the username based on the provided rules."""
        if not entered_username.strip():
            return False, "Username cannot be blank."
        
        if entered_username.isdigit():
            return False, "Username cannot be all numbers."
        
        if "@" in entered_username and "." in entered_username:
            return False, "Username cannot be an email address."
        
        return True, "Valid username."
    
    def submit_username(event=None):
        print(f'entered submit username')
        entered_username = entry.get()
        valid, message = is_valid_username(entered_username)
        
        if valid:
            nonlocal username  # Declare nonlocal to update username within the nested function
            username = entered_username  # Store the valid username
            root.quit()  # Stop the local mainloop
            root.destroy()  # Close the username entry window
        else:
            messagebox.showerror("Error", message)
    
    # Set up the username variable to hold the result
    username = None
    
    # Initialize the main window
    ctk.set_appearance_mode("dark")  # Modes: "dark" or "light"
    ctk.set_default_color_theme("blue")  # We will override the default colors manually
    
    root = ctk.CTk()
    root.title("Username Entry")

    
    # Override the color theme to use the clemson orange
    orange_color = "#F56600"
    # Force geometry to screen size to ensure full coverage
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    root.geometry(f"{screen_width}x{screen_height}+0+0")
    
    #Create a central frame for centering elements
    frame = ctk.CTkFrame(master=root)
    frame.pack(expand=True)

    # Title label
    title_label = ctk.CTkLabel(master=frame, text="Welcome to the Makerspace! Enter Your Clemson Username:", font=("Arial", 55), text_color=orange_color)
    title_label.pack(pady=20)

    # Instruction label
    label = ctk.CTkLabel(master=frame, text="The part before the @clemson.edu", font=("Arial", 40))
    label.pack(pady=10)
    
    # Create an entry box with focus on open
    entry = ctk.CTkEntry(master=frame, width=500, height=50, placeholder_text="Enter username", font=("Arial", 32))
    entry.pack(pady=10)

    # Bind the Enter key to submit the form
    root.bind('<Return>', submit_username)  # Bind Enter (Return) key to the submit function
    
    root.mainloop() # Start the application loop
    #time.sleep(3)
    #entry.focus_set()  # Automatically focus on the entry box
    #keyboard.press(Key.tab)
    #print(f'tab pressed')
    # Return the username after the window closes
    return username

def main():
    userstatus=None
    hardware_id = sys.argv[1] #This gets the hardware ID from the gloabl system variables as defined from the other script to pass along the variables.
    workbook,sheet,sheet2 = load_excel()
    username = find_hardware_id(sheet2, hardware_id)
    first_name=None
    major=None
    root = tk.Tk()
    root.withdraw()  # Hide the root window initially
    root.bind("<Escape>", close_on_escape) # Bind the Escape key to close the program

    if username != None:
        print(f"User found: {username}")
        userstatus=0
        first_name,last_name,major = find_userdata(hardware_id, sheet2)
        add_user_to_sheet(sheet_name,sheet2_name,hardware_id, username,first_name,last_name,major,workbook,userstatus)
        show_welcome_popup(root,username,first_name,userstatus)
        #api_call(hardware_id,Location,username)
        root.deiconify()  # Show the window
        make_fullscreen_on_top(root)
        root.mainloop()
    else:
        print("New user detected. Prompting for username.")
        userstatus=1
        username = prompt_for_username()
        print(f'Username entered: {username}')
        if username != None:
            first_name, last_name, major = scrape_user(username)
            show_welcome_popup(root, username, first_name, userstatus)
            root.deiconify()  # Show the window
            print(f'root deiconified')
            make_fullscreen_on_top(root)
            add_user_to_sheet(sheet_name,sheet2_name,hardware_id, username,first_name,last_name,major,workbook,userstatus)
            workbook.save(file_path)
            #api_call(hardware_id,Location,username)
        else:
            print('Username Prompt timed out')
        username=None    
            
    
if __name__ == "__main__":
    main()