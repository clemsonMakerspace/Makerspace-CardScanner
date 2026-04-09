"""
Fetch Missing Training Data for All Users
==========================================
Scans the database and Excel sheet for users with no training data,
then calls the Bridge LMS API to populate it for each one.
Runs a full bidirectional sync when complete.

Run via:   fetch_missing_training.bat
       or: python fetch_missing_training.py [--dry-run] [--excel-only] [--db-only]
"""

import sys as _sys
if hasattr(_sys.stdout, 'reconfigure'):
    _sys.stdout.reconfigure(encoding='utf-8', errors='replace')
del _sys



import sys
import os
import json
import time
import argparse
from datetime import datetime

# Ensure the script finds modules in its own directory
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# ---------------------------------------------------------------------------
# Imports (fail loudly with helpful messages)
# ---------------------------------------------------------------------------
try:
    import database as db
except ImportError:
    print("ERROR: database.py not found. Run this script from the project root.")
    sys.exit(1)

try:
    from bridge_api import get_all_training_status, API_CONFIGURED
except ImportError:
    print("ERROR: bridge_api.py not found.")
    sys.exit(1)

try:
    from excel_db_sync import sync_excel_to_database, sync_database_to_excel
except ImportError:
    print("ERROR: excel_db_sync.py not found.")
    sys.exit(1)

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

EXCEL_FILE = "hardware_users.xlsx"
USERS_SHEET = "Users"
DB_FILE = "hardware_users.db"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def divider(char="=", width=70):
    print(char * width)


def ensure_database_exists():
    """
    Check that the SQLite database file exists.  If it is missing (common
    when porting an older Excel-only installation), create it from scratch
    and populate it from the Excel sheet so the rest of the script has a
    complete database to work with.

    Returns True if the database was freshly created, False if it already existed.
    """
    if os.path.exists(DB_FILE):
        return False

    print("  Database file not found -- creating a new one...")
    db.init_database()

    if not os.path.exists(EXCEL_FILE):
        print("  Excel file also missing -- empty database created.")
        print("  Place hardware_users.xlsx next to this script and re-run.")
        return True

    print(f"  Importing users and scans from {EXCEL_FILE}...")
    sync_excel_to_database()
    stats = db.get_database_stats()
    print(f"  Database created: {stats.get('total_users', 0)} users, "
          f"{stats.get('total_scans', 0)} scans imported.")
    return True


def get_users_missing_training_from_db():
    """
    Return list of dicts for every user in the SQLite DB that has no
    training_data (NULL or empty string).
    """
    missing = []
    try:
        with db.get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT username, hardware_id, first_name, last_name
                FROM users
                WHERE training_data IS NULL OR training_data = ''
                ORDER BY username
            """)
            for row in cursor.fetchall():
                missing.append(dict(row))
    except Exception as exc:
        print(f"  ERROR querying database: {exc}")
    return missing


def get_users_missing_training_from_excel():
    """
    Return list of dicts for every user row in the Excel Users sheet that
    has no training data in column G (index 6).
    """
    missing = []
    if not OPENPYXL_AVAILABLE:
        print("  WARNING: openpyxl not available, skipping Excel scan.")
        return missing

    if not os.path.exists(EXCEL_FILE):
        print(f"  WARNING: {EXCEL_FILE} not found, skipping Excel scan.")
        return missing

    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb[USERS_SHEET]
        for row in ws.iter_rows(min_row=2, values_only=True):
            username = row[0] if len(row) > 0 else None
            hardware_id = row[1] if len(row) > 1 else None
            first_name = row[3] if len(row) > 3 else None
            last_name = row[4] if len(row) > 4 else None
            training_json = row[6] if len(row) > 6 else None

            if not username:
                continue

            if not training_json:
                missing.append({
                    "username": username,
                    "hardware_id": hardware_id,
                    "first_name": first_name,
                    "last_name": last_name,
                    "source": "excel"
                })
        wb.close()
    except Exception as exc:
        print(f"  ERROR reading Excel file: {exc}")

    return missing


def merge_missing_users(from_db, from_excel):
    """
    Combine both lists, deduplicate by username (case-insensitive).
    Marks source so the user knows where the gap came from.
    """
    seen = {}
    for u in from_db:
        key = u["username"].lower()
        u.setdefault("source", "db")
        seen[key] = u
    for u in from_excel:
        key = u["username"].lower()
        if key not in seen:
            seen[key] = u
        else:
            # Already in list from DB; note both have gaps
            seen[key]["source"] = "db+excel"
    return list(seen.values())


def fetch_and_save(username, dry_run=False):
    """
    Fetch training status from Bridge API and persist to DB.
    Returns (success: bool, info: str)
    """
    if dry_run:
        return True, "[dry-run] skipped API call"

    try:
        status = get_all_training_status(username)
        if status is None:
            return False, "API returned None (check credentials/network)"

        # Save to database
        saved = db.update_training_data(username, status)
        if not saved:
            # User may not be in DB yet (Excel-only user) -- insert them first
            hw = None
            fn = None
            ln = None
            try:
                with db.get_db_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute(
                        "SELECT hardware_id, first_name, last_name FROM users WHERE username=?",
                        (username,)
                    )
                    row = cursor.fetchone()
                    if row:
                        hw, fn, ln = row
            except Exception:
                pass

            if hw is not None:
                db.add_or_update_user(username, hw, fn, ln, None)
                db.update_training_data(username, status)

        completed = status.get("completed_courses", 0)
        total = status.get("total_courses", 0)
        required_ok = status.get("required_complete", False)
        tag = " [REQ-COMPLETE]" if required_ok else ""
        return True, f"{completed}/{total} courses complete{tag}"

    except Exception as exc:
        return False, str(exc)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Fetch Bridge LMS training data for users missing it."
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Find missing users and report totals but do NOT call the API."
    )
    parser.add_argument(
        "--db-only", action="store_true",
        help="Only scan the SQLite database (skip Excel sheet)."
    )
    parser.add_argument(
        "--excel-only", action="store_true",
        help="Only scan the Excel sheet (skip SQLite database)."
    )
    parser.add_argument(
        "--no-sync", action="store_true",
        help="Skip the final database<->Excel sync step."
    )
    parser.add_argument(
        "--delay", type=float, default=0.3,
        help="Seconds to wait between API calls (default 0.3, set 0 to disable)."
    )
    args = parser.parse_args()

    divider()
    print("  MAKERSPACE -- FETCH MISSING TRAINING DATA")
    divider()
    print(f"  Started : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    if args.dry_run:
        print("  Mode    : DRY RUN (no API calls, no data written)")
    print()

    # --- Guard: API must be configured ------------------------------------
    if not API_CONFIGURED and not args.dry_run:
        print("ERROR: Bridge API is not configured.")
        print("  Open config.py and set BRIDGE_API_URL and BRIDGE_AUTH_TOKEN.")
        sys.exit(1)

    # --- Step 0: Ensure the database file exists ---------------------------
    #     (Handles first-run on machines that only have the Excel sheet.)
    print("Step 0/4  Checking for database file...")
    db_was_created = ensure_database_exists()
    if db_was_created:
        print("  New database created and populated from Excel.")
    else:
        print("  Database found.")
    print()

    # --- Step 1: Sync Excel -> DB so the DB has the latest manual edits ----
    if not args.no_sync:
        print("Step 1/4  Syncing Excel -> Database (capturing manual edits)...")
        sync_excel_to_database()
    else:
        print("Step 1/4  Skipped (--no-sync).")
    print()

    # --- Step 2: Collect users missing training data -----------------------
    print("Step 2/4  Finding users with missing training data...")

    from_db = [] if args.excel_only else get_users_missing_training_from_db()
    from_excel = [] if args.db_only else get_users_missing_training_from_excel()

    db_missing_count = len(from_db)
    excel_missing_count = len(from_excel)

    all_missing = merge_missing_users(from_db, from_excel)
    total = len(all_missing)

    print(f"  Database : {db_missing_count} users with no training data")
    print(f"  Excel    : {excel_missing_count} users with no training data")
    print(f"  Combined : {total} unique users to update")
    print()

    if total == 0:
        print("Nothing to do -- all users already have training data.")
    else:
        # --- Step 3: Fetch training for each user --------------------------
        print(f"Step 3/4  Fetching training data ({total} users)...")
        if args.dry_run:
            print("  [dry-run] Listing users that would be updated:")

        divider("-")

        success_count = 0
        fail_count = 0
        fail_list = []

        for i, user in enumerate(all_missing, 1):
            username = user["username"]
            name_parts = [user.get("first_name") or "", user.get("last_name") or ""]
            display_name = " ".join(p for p in name_parts if p).strip() or "(no name)"
            source_tag = f"[{user.get('source', '?')}]"

            prefix = f"  [{i:>4}/{total}]"
            print(f"{prefix} {username:<16} {display_name:<30} {source_tag}", end="  ", flush=True)

            ok, info = fetch_and_save(username, dry_run=args.dry_run)

            if ok:
                print(f"OK  {info}")
                success_count += 1
            else:
                print(f"FAIL  {info}")
                fail_count += 1
                fail_list.append((username, info))

            # Rate-limit to be kind to the API
            if args.delay > 0 and i < total:
                time.sleep(args.delay)

        divider("-")
        print()
        print(f"  Updated : {success_count}")
        print(f"  Failed  : {fail_count}")
        if fail_list:
            print("  Failures:")
            for un, reason in fail_list:
                print(f"    {un}: {reason}")
        print()

    # --- Step 4: Sync DB -> Excel ------------------------------------------
    if not args.no_sync:
        print("Step 4/4  Syncing Database -> Excel (writing updated training data)...")
        sync_database_to_excel()
    else:
        print("Step 4/4  Skipped (--no-sync).")

    print()
    divider()
    print(f"  Completed : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    divider()


if __name__ == "__main__":
    main()
