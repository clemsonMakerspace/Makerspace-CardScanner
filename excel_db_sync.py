"""
Bidirectional Excel ↔ Database Synchronization
Syncs data between Excel and SQLite database on startup/shutdown
Preserves manual edits to Excel while maintaining database performance
"""

import sqlite3
import json
from datetime import datetime
from openpyxl import load_workbook
import database as db
from excel_utils import safe_excel_read, safe_excel_write

EXCEL_FILE = "hardware_users.xlsx"
USERS_SHEET = "Users"
SCANS_SHEET = "Scans"


def get_excel_modification_time():
    """Get last modification time of Excel file"""
    import os
    if os.path.exists(EXCEL_FILE):
        return os.path.getmtime(EXCEL_FILE)
    return 0


def get_database_modification_time():
    """Get last modification time of database file"""
    import os
    if os.path.exists(db.DB_FILE):
        return os.path.getmtime(db.DB_FILE)
    return 0


def sync_excel_to_database():
    """
    Import data from Excel to Database
    This allows users to manually edit Excel and have changes sync to DB
    """
    print("\n" + "="*60)
    print("📊 Syncing Excel → Database...")
    print("="*60)
    
    try:
        with safe_excel_read(EXCEL_FILE) as wb:
            users_sheet = wb[USERS_SHEET]
            scans_sheet = wb[SCANS_SHEET]
            
            # Sync Users
            users_synced = 0
            for row in users_sheet.iter_rows(min_row=2, values_only=True):
                username = row[0]  # Column A
                hardware_id = row[1]  # Column B
                login_count = row[2] if len(row) > 2 else 0  # Column C
                first_name = row[3] if len(row) > 3 else None  # Column D
                last_name = row[4] if len(row) > 4 else None  # Column E
                major = row[5] if len(row) > 5 else None  # Column F
                training_json = row[6] if len(row) > 6 else None  # Column G
                training_updated = row[7] if len(row) > 7 else None  # Column H
                
                # Skip empty rows
                if not username and not hardware_id:
                    continue
                
                # Add/update user in database
                if username and hardware_id:
                    db.add_or_update_user(username, hardware_id, first_name, last_name, major)
                    
                    # Update training data if present
                    if training_json:
                        try:
                            training_data = json.loads(training_json)
                            db.update_training_data(username, training_data)
                        except (json.JSONDecodeError, TypeError):
                            pass  # Skip invalid JSON
                    
                    users_synced += 1
            
            # Sync Scans (only new ones not already in database)
            scans_synced = 0
            for row in scans_sheet.iter_rows(min_row=2, values_only=True):
                hardware_id = row[0] if len(row) > 0 else None  # Column A
                username = row[1] if len(row) > 1 else ""  # Column B
                timestamp = row[2] if len(row) > 2 else None  # Column C
                location = row[3] if len(row) > 3 else "Watt"  # Column D
                
                # Skip empty rows
                if not hardware_id and not timestamp:
                    continue
                
                # Convert timestamp to string if it's a datetime object
                if timestamp:
                    if hasattr(timestamp, 'strftime'):
                        timestamp_str = timestamp.strftime('%m/%d/%Y %H:%M:%S')
                    else:
                        timestamp_str = str(timestamp)
                    
                    # Check if scan already exists (avoid duplicates)
                    with db.get_db_connection() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            SELECT COUNT(*) FROM scans
                            WHERE hardware_id = ? AND timestamp = ?
                        ''', (hardware_id, timestamp_str))
                        exists = cursor.fetchone()[0] > 0

                        if not exists:
                            # Insert directly using the open connection instead of
                            # calling db.add_scan(), which would deadlock trying to
                            # re-acquire the non-reentrant _db_lock we already hold.
                            cursor.execute('''
                                INSERT INTO scans (hardware_id, username, timestamp, location)
                                VALUES (?, ?, ?, ?)
                            ''', (hardware_id, username, timestamp_str, location))
                            cursor.execute('''
                                UPDATE users
                                SET login_count = login_count + 1,
                                    updated_at = CURRENT_TIMESTAMP
                                WHERE username = ?
                            ''', (username,))
                            scans_synced += 1
            
            print(f"✓ Synced {users_synced} users from Excel to Database")
            print(f"✓ Synced {scans_synced} new scans from Excel to Database")
            print("="*60 + "\n")
            return True
            
    except Exception as e:
        print(f"✗ Error syncing Excel to Database: {e}")
        print("="*60 + "\n")
        return False


def sync_database_to_excel():
    """
    Export data from Database to Excel
    This preserves database changes and updates Excel with latest data
    """
    print("\n" + "="*60)
    print("💾 Syncing Database → Excel...")
    print("="*60)
    
    try:
        with safe_excel_write(EXCEL_FILE) as wb:
            users_sheet = wb[USERS_SHEET]
            scans_sheet = wb[SCANS_SHEET]
            
            # Clear existing data (keep headers)
            users_sheet.delete_rows(2, users_sheet.max_row)
            scans_sheet.delete_rows(2, scans_sheet.max_row)
            
            # Sync Users from database
            with db.get_db_connection() as conn:
                cursor = conn.cursor()
                
                # Get all users
                cursor.execute('''
                    SELECT username, hardware_id, login_count, first_name, last_name, 
                           major, training_data, training_last_updated
                    FROM users
                    ORDER BY username
                ''')
                
                users_synced = 0
                for row in cursor.fetchall():
                    users_sheet.append([
                        row[0],  # username
                        row[1],  # hardware_id
                        row[2] or 0,  # login_count
                        row[3],  # first_name
                        row[4],  # last_name
                        row[5],  # major
                        row[6],  # training_data (JSON)
                        row[7]   # training_last_updated
                    ])
                    users_synced += 1
                
                # Get all scans
                cursor.execute('''
                    SELECT hardware_id, username, timestamp, location
                    FROM scans
                    ORDER BY timestamp DESC
                ''')
                
                scans_synced = 0
                for row in cursor.fetchall():
                    scans_sheet.append([
                        row[0],  # hardware_id
                        row[1],  # username
                        row[2],  # timestamp
                        row[3]   # location
                    ])
                    scans_synced += 1
                
                print(f"✓ Synced {users_synced} users from Database to Excel")
                print(f"✓ Synced {scans_synced} scans from Database to Excel")
                print("="*60 + "\n")
                return True
                
    except Exception as e:
        print(f"✗ Error syncing Database to Excel: {e}")
        print("="*60 + "\n")
        return False


def smart_bidirectional_sync():
    """
    Intelligent bidirectional sync:
    - Compares Excel vs Database modification times
    - If Excel is newer: Excel → Database (user made manual edits)
    - If Database is newer: Database → Excel (program updated database)
    - Then syncs in both directions to ensure consistency
    """
    print("\n" + "🔄 " + "="*58)
    print("   SMART BIDIRECTIONAL SYNC - Excel ↔ Database")
    print("🔄 " + "="*58)
    
    excel_time = get_excel_modification_time()
    db_time = get_database_modification_time()
    
    if excel_time == 0 and db_time == 0:
        print("⚠ No Excel or Database found - will create on first use")
        print("="*60 + "\n")
        return
    
    if excel_time == 0:
        print("📊 Excel file missing - creating from database...")
        sync_database_to_excel()
        return
    
    if db_time == 0:
        print("💾 Database missing - creating from Excel...")
        db.init_database()
        sync_excel_to_database()
        return
    
    # Both exist - determine which is newer
    excel_date = datetime.fromtimestamp(excel_time).strftime('%Y-%m-%d %H:%M:%S')
    db_date = datetime.fromtimestamp(db_time).strftime('%Y-%m-%d %H:%M:%S')
    
    print(f"📊 Excel modified:    {excel_date}")
    print(f"💾 Database modified: {db_date}")
    
    time_diff = abs(excel_time - db_time)
    
    if time_diff < 5:  # Less than 5 seconds difference - consider them in sync
        print("✓ Already in sync (modified within 5 seconds)")
        print("="*60 + "\n")
        return
    
    if excel_time > db_time:
        print("→ Excel is newer - User likely made manual edits")
        print("  Priority: Excel → Database")
        sync_excel_to_database()
        # Then update Excel with any database-only data
        sync_database_to_excel()
    else:
        print("→ Database is newer - Program made updates")
        print("  Priority: Database → Excel")
        sync_database_to_excel()
        # Then import any Excel-only data
        sync_excel_to_database()
    
    print("✅ Bidirectional sync complete!")
    print("="*60 + "\n")


def sync_on_startup():
    """Run bidirectional sync when program starts"""
    print("\n🚀 STARTUP SYNC")
    smart_bidirectional_sync()


def sync_on_shutdown():
    """Run database → Excel sync when program closes"""
    print("\n🛑 SHUTDOWN SYNC")
    print("💾 Ensuring Excel has latest database changes...")
    sync_database_to_excel()


if __name__ == "__main__":
    # Test the sync system
    print("Manual Bidirectional Sync Test")
    smart_bidirectional_sync()
