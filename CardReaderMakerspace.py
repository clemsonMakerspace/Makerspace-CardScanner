import sys
import ctypes

# Set DPI awareness ONCE at process startup, before any GUI toolkit imports.
# This prevents customtkinter's CTk() from changing the DPI context mid-process,
# which would cause subsequent tkinter windows to report wrong screen dimensions.
try:
    ctypes.windll.shcore.SetProcessDpiAwareness(2)  # PROCESS_PER_MONITOR_DPI_AWARE
except Exception:
    pass

import customtkinter as ctk
from openpyxl import load_workbook
import tkinter as tk
import PIL
from PIL import Image, ImageTk 
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import threading
import os
import json
import time  # TEMP DEBUG: For performance timing
import atexit  # For cleanup on program exit
import subprocess  # For launching queued scans

# Import Excel utilities for safe file operations
from excel_utils import safe_excel_write, safe_excel_read, create_backup, verify_excel_integrity

# Import SQLite database module for robust backup
import database as db

# Import Excel ↔ Database bidirectional sync
try:
    from excel_db_sync import sync_on_startup, sync_on_shutdown
    BIDIRECTIONAL_SYNC_AVAILABLE = True
except ImportError:
    print("Warning: excel_db_sync module not found. Bidirectional sync disabled.")
    BIDIRECTIONAL_SYNC_AVAILABLE = False

# Import database sync (optional, based on config)
try:
    from database_sync import init_sync
    SYNC_AVAILABLE = True
except ImportError:
    SYNC_AVAILABLE = False

# Import Bridge API for training status
try:
    from bridge_api import get_all_training_status
    BRIDGE_API_AVAILABLE = True
except ImportError:
    print("Warning: bridge_api module not found. Training status will not be displayed.")
    BRIDGE_API_AVAILABLE = False

# Import minimal training display setting from config
try:
    from config import MINIMAL_TRAINING_DISPLAY
except ImportError:
    MINIMAL_TRAINING_DISPLAY = False

def get_resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


###    Version 1.2

# Path to excel sheet
file_path = "hardware_users.xlsx"
sheet_name = "Scans"
sheet2_name = "Users"
Location = "Watt"

# OPTIMIZATION: Cache for background image to avoid repeated loading
_cached_background_image = None  # Cache the loaded background image
_cached_screen_size = None  # Cache screen dimensions

# Queue system for handling scans during popup display
POPUP_FLAG_FILE = ".popup_active"  # File flag to track popup across processes
QUEUED_SCAN_FILE = ".queued_scan"  # File to store queued hardware ID
POPUP_TIMEOUT = 10  # Maximum popup lifetime in seconds (safety timeout)

# Global variable for tracking hardware ID scanned during username entry
pending_hardware_id = None


def cleanup_stale_flags():
    """Remove stale flag files on startup to prevent orphaned locks"""
    try:
        if os.path.exists(POPUP_FLAG_FILE):
            # Check if flag is stale (older than popup timeout)
            flag_age = time.time() - os.path.getmtime(POPUP_FLAG_FILE)
            if flag_age > POPUP_TIMEOUT:
                os.remove(POPUP_FLAG_FILE)
                print(f"[CLEANUP] Removed stale popup flag ({flag_age:.1f}s old)")
            else:
                print(f"[CLEANUP] Active popup flag detected ({flag_age:.1f}s old)")
        
        if os.path.exists(QUEUED_SCAN_FILE):
            # Queued scans older than 30 seconds are likely orphaned
            scan_age = time.time() - os.path.getmtime(QUEUED_SCAN_FILE)
            if scan_age > 30:
                os.remove(QUEUED_SCAN_FILE)
                print(f"[CLEANUP] Removed stale queued scan ({scan_age:.1f}s old)")
    except Exception as e:
        print(f"[CLEANUP] Error during cleanup: {e}")


def is_popup_active():
    """Check if a popup is currently displayed (across processes)"""
    if os.path.exists(POPUP_FLAG_FILE):
        # Additional safety check: ignore flags older than timeout
        try:
            flag_age = time.time() - os.path.getmtime(POPUP_FLAG_FILE)
            if flag_age > POPUP_TIMEOUT:
                print(f"[SAFETY] Ignoring stale popup flag ({flag_age:.1f}s old)")
                set_popup_active(False)  # Clean it up
                return False
            return True
        except:
            return True  # If we can't check age, assume active to be safe
    return False


def set_popup_active(active=True):
    """Set popup active state (across processes)"""
    if active:
        # Create flag file
        with open(POPUP_FLAG_FILE, 'w') as f:
            f.write(str(datetime.now()))
    else:
        # Remove flag file
        if os.path.exists(POPUP_FLAG_FILE):
            try:
                os.remove(POPUP_FLAG_FILE)
            except:
                pass


def queue_scan(hardware_id):
    """Queue a scan for later processing"""
    with open(QUEUED_SCAN_FILE, 'w') as f:
        f.write(str(hardware_id))
    print(f"[QUEUE] Scan queued: {hardware_id}")


def get_queued_scan():
    """Get queued scan and clear the queue"""
    if os.path.exists(QUEUED_SCAN_FILE):
        try:
            with open(QUEUED_SCAN_FILE, 'r') as f:
                hardware_id = f.read().strip()
            os.remove(QUEUED_SCAN_FILE)
            return hardware_id
        except:
            return None
    return None


def load_excel():
    """Load the workbook and sheet with error handling"""
    try:
        # Verify file integrity before loading
        if not verify_excel_integrity(file_path):
            print("WARNING: Excel file may be corrupted! Attempting to use it anyway...")
        
        workbook = load_workbook(filename=file_path)
        sheet = workbook[sheet_name]
        sheet2 = workbook[sheet2_name]
        return workbook, sheet, sheet2
    except Exception as e:
        print(f"ERROR loading Excel file: {e}")
        raise

def find_hardware_id(sheet2, hardware_id):
    """
    Fast lookup: Check database first (indexed), then Excel as fallback
    """
    # Try database first (much faster with index on hardware_id)
    try:
        user = db.get_user_by_hardware_id(hardware_id)
        if user:
            return user['username']
    except Exception as e:
        print(f"Database lookup failed, falling back to Excel: {e}")
    
    # Fallback to Excel if database fails
    for row in sheet2.iter_rows(min_row=2, max_row=sheet2.max_row, values_only=True):
        if str(row[1]) == str(hardware_id):  # Compare hardware_id in column B (index 1)
            return row[0]  # Return the username from column A (index 0)
    return None  # Return None if not found

def find_userdata(hardware_id, sheet2):
    """
    Fast lookup: Check database first (indexed), then Excel as fallback
    """
    # Try database first (much faster with index on hardware_id)
    try:
        user = db.get_user_by_hardware_id(hardware_id)
        if user:
            return user['first_name'], user['last_name'], user['major']
    except Exception as e:
        print(f"Database lookup failed, falling back to Excel: {e}")
    
    # Fallback to Excel if database fails
    for row in sheet2.iter_rows(min_row=2, max_row=sheet2.max_row, values_only=True):
        if str(row[1]) == str(hardware_id):  # Look for hardware_id in col. B
            first_name = row[3]  # Column D 
            last_name = row[4]   # Column E 
            major = row[5]       # Column F
            return first_name, last_name, major
    return None, None, None  # Set to None if they are not found

def get_training_data_from_excel(username, sheet2):
    """
    Fast lookup: Check database first (indexed), then Excel as fallback
    Returns the training status dict or None if not found/invalid
    """
    # Try database first (much faster with index on username)
    try:
        training_data = db.get_training_data(username)
        if training_data:
            print(f"Loaded training data from database for {username}")
            return training_data
    except Exception as e:
        print(f"Database training lookup failed, falling back to Excel: {e}")
    
    # Fallback to Excel if database fails
    for row in sheet2.iter_rows(min_row=2, max_row=sheet2.max_row, values_only=False):
        if row[0].value and str(row[0].value).lower() == username.lower():
            # Column G (index 6) stores training JSON data
            # Column H (index 7) stores last API update timestamp
            training_json = row[6].value if len(row) > 6 else None
            last_updated = row[7].value if len(row) > 7 else None
            
            if training_json:
                try:
                    training_data = json.loads(training_json)
                    training_data['last_excel_update'] = last_updated
                    print(f"Loaded training data from Excel for {username} (updated: {last_updated})")
                    return training_data
                except (json.JSONDecodeError, TypeError) as e:
                    print(f"Error parsing training JSON for {username}: {e}")
                    return None
            break
    return None

def save_training_data_to_excel(username, training_status):
    """
    Save training data to both Excel and SQLite database
    Column G: Training data as JSON
    Column H: Last update timestamp
    """
    if not training_status:
        return
    
    # Save to database first (more reliable)
    db.update_training_data(username, training_status)
    
    # Then save to Excel
    try:
        with safe_excel_write(file_path) as wb:
            users_sheet = wb[sheet2_name]
            
            # Find the user's row
            for row in users_sheet.iter_rows(min_row=2, values_only=False):
                if row[0].value and str(row[0].value).lower() == username.lower():
                    # Ensure columns G and H exist
                    if len(row) < 8:
                        print(f"Warning: Row doesn't have enough columns for {username}")
                        return
                    
                    # Save as compact JSON
                    training_json = json.dumps(training_status, separators=(',', ':'))
                    timestamp = datetime.now().strftime('%m/%d/%Y %H:%M:%S')
                    
                    row[6].value = training_json  # Column G
                    row[7].value = timestamp       # Column H
                    
                    print(f"Training data saved to Excel for {username} at {timestamp}")
                    break
    except Exception as e:
        print(f"Error saving training data to Excel: {e}")
        # Database save still succeeded, so don't fail completely

def update_training_data_async(username):
    """
    Fetch training data from API and update Excel in background thread
    This doesn't block the main UI
    """
    def background_update():
        if BRIDGE_API_AVAILABLE:
            try:
                bg_start = time.perf_counter()  # API DEBUG
                print(f"[Background] Fetching training data for {username}...")
                status = get_all_training_status(username)
                if status:
                    save_training_data_to_excel(username, status)
                    bg_total = (time.perf_counter() - bg_start) * 1000  # API DEBUG
                    print(f"[Background] Training data updated for {username}")
                    print(f"[API DEBUG] Background update total time: {bg_total:.0f}ms")  # API DEBUG
                else:
                    bg_total = (time.perf_counter() - bg_start) * 1000  # API DEBUG
                    print(f"[Background] No training data returned for {username}")
                    print(f"[API DEBUG] Background update (no data): {bg_total:.0f}ms")  # API DEBUG
            except Exception as e:
                bg_total = (time.perf_counter() - bg_start) * 1000  # API DEBUG
                print(f"[API DEBUG] Background update (error): {bg_total:.0f}ms")  # API DEBUG
                print(f"[Background] Error updating training data for {username}: {e}")
    
    thread = threading.Thread(target=background_update)
    thread.start()
    print(f"Started background training data update for {username}")
    return thread


def add_user_to_sheet(sheet_name, sheet2_name, hardware_id, username, first_name, last_name, major, workbook, userstatus):
    """Add user and scan data to both Excel and SQLite database"""
    
    save_start = time.perf_counter()  # TEMP DEBUG
    
    # Add to database first (more reliable and FAST)
    if userstatus == 1:
        # New user
        db.add_or_update_user(username, hardware_id, first_name, last_name, major)
    
    # Add scan to database (FAST - ~1ms)
    now = datetime.now()
    timestamp = now.strftime('%m/%d/%Y %H:%M:%S')
    db.add_scan(hardware_id, username, timestamp, Location)
    
    db_time = (time.perf_counter() - save_start) * 1000  # TEMP DEBUG
    print(f"[DEBUG] Database save: {db_time:.2f}ms")  # TEMP DEBUG
    
    # OPTIMIZATION: Skip Excel write for existing users (database is source of truth)
    # Only write to Excel for new users or critical updates
    if userstatus == 1:
        # New user - must update Excel
        excel_start = time.perf_counter()  # TEMP DEBUG
        try:
            with safe_excel_write(file_path) as wb:
                scans_sheet = wb[sheet_name]
                users_sheet = wb[sheet2_name]
                
                # Search for matching hardware ID in the "Users" sheet or for an empty hardware ID cell
                for row in users_sheet.iter_rows(min_row=2, values_only=False):  # Skip header row
                    cell_hardware_id = row[1].value  # Column B in "Users" for hardware_id

                    # Cast both the hardware_id from input and the one from the sheet to str for comparison
                    if str(cell_hardware_id) == str(hardware_id):
                        print(f"User with hardware ID {hardware_id} already exists in 'Users' sheet.")
                        break  # Stop searching after finding the match

                    # If the hardware ID cell is empty (i.e., new entry row), fill in this row
                    if cell_hardware_id is None or cell_hardware_id == "":
                        row[0].value = username  # Column A for username
                        row[1].value = int(hardware_id)  # Column B for hardware ID
                        row[3].value = first_name  # Column D for first name
                        row[4].value = last_name   # Column E for last name
                        row[5].value = major       # Column F for major
                        print(f"New user {first_name} {last_name} added to 'Users' sheet.")
                        break  # Stop searching after appending the new data

                # Add the scan to the 'Scans' sheet
                scans_sheet.append([int(hardware_id), username, timestamp])
                
            excel_time = (time.perf_counter() - excel_start) * 1000  # TEMP DEBUG
            print(f"[DEBUG] Excel write (new user): {excel_time:.2f}ms")  # TEMP DEBUG
            print(f"New user added to Excel and database.")
        except Exception as e:
            print(f"ERROR: Failed to save to Excel: {e}")
            print(f"Note: Data was saved to database successfully")
    else:
        # OPTIMIZATION: Existing user - skip Excel write, database is enough
        # Excel can be synced periodically in background
        print(f"[DEBUG] Excel write: SKIPPED (existing user, database updated)")  # TEMP DEBUG
        print(f"Scan added to database.")
        print("Note: Data was saved to database successfully")
        # Don't raise - allow program to continue even if Excel write fails

def scrape_user(username):
    scrape_start = time.perf_counter()  # TEMP DEBUG
    print(f"[DEBUG] Starting web scrape for {username}...")  # TEMP DEBUG
    
    # Set up Selenium with headless Chrome
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Run Chrome in headless mode
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    # Suppress USB, GCM, and DevTools logging noise
    chrome_options.add_argument("--disable-usb-discovery")
    chrome_options.add_argument("--disable-features=MediaRouter")
    chrome_options.add_argument("--log-level=3")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])

    # Set up the driver
    driver_start = time.perf_counter()  # TEMP DEBUG
    driver = webdriver.Chrome(options=chrome_options)
    driver_time = (time.perf_counter() - driver_start) * 1000  # TEMP DEBUG
    print(f"[DEBUG] Chrome driver init: {driver_time:.2f}ms")  # TEMP DEBUG

    # Create the URL using username
    url = f"https://my.clemson.edu/#/directory/person/{username}"

    # Load the page
    page_start = time.perf_counter()  # TEMP DEBUG
    driver.get(url)
    page_time = (time.perf_counter() - page_start) * 1000  # TEMP DEBUG
    print(f"[DEBUG] Page load: {page_time:.2f}ms")  # TEMP DEBUG

    # Wait for the full name element to appear
    try:
        WebDriverWait(driver, 1).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '.personView .primaryInfo h2'))
        )
    except TimeoutException:
        print(f"Timeout while waiting for the element on the page for {username}.")
        driver.quit()
        return None, None, None
    except Exception as e:
        print(f"Error during page load: {e}")
        driver.quit()
        return None, None, None

    # Get the page
    page_source = driver.page_source

    # Parse the page with BeautifulSoup
    soup = BeautifulSoup(page_source, 'html.parser')

    # Find the Name
    full_name_element = soup.select_one('.personView .primaryInfo h2')
    if full_name_element:
        full_name = full_name_element.get_text().strip()
        name_parts = full_name.split()
        first_name = name_parts[0]  # The first part of the name
        last_name = name_parts[-1]  # The last part of the name
    else:
        first_name, last_name = None, None

    # Find major
    major_element = soup.select_one('.personView .primaryInfo .data p')
    major = major_element.get_text().strip() if major_element else None

    # Print the scraped information
    print(f"First Name: {first_name}")
    print(f"Last Name: {last_name}")
    print(f"Major: {major}")

    # Close the driver
    driver.quit()
    
    total_scrape_time = (time.perf_counter() - scrape_start) * 1000  # TEMP DEBUG
    print(f"[DEBUG] TOTAL SCRAPE TIME: {total_scrape_time:.2f}ms")  # TEMP DEBUG
    
    return first_name, last_name, major

def make_fullscreen_on_top(root):
    root.attributes('-fullscreen', True)
    root.attributes('-topmost', True)

def show_welcome_popup(root, username, first_name, userstatus, training_status=None):
    set_popup_active(True)  # Mark popup as active (file flag)
    
    try:
        popup_create_start = time.perf_counter()  # TEMP DEBUG
        print(f"[DEBUG] Creating popup window...")  # TEMP DEBUG
        
        # Ensure fullscreen is applied and updated before measuring
        root.attributes('-fullscreen', True)
        root.attributes('-topmost', True)
        root.update_idletasks()
        
        # Get screen dimensions from tkinter (consistent with its own rendering)
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        print(f"[DEBUG] Screen dimensions: {screen_width}x{screen_height}")  # TEMP DEBUG
        
        # Load and scale background image (location-specific)
        img_start = time.perf_counter()  # TEMP DEBUG
        print(f"[DEBUG] Loading background image...")  # TEMP DEBUG
        
        if Location == "Cooper":
            background_filename = "BackgroundAdobe.png"
        else:
            background_filename = "BackgroundWatt.png"
        
        image = Image.open(get_resource_path(background_filename))
        image = image.resize((screen_width, screen_height), PIL.Image.Resampling.LANCZOS)
        bg_image = ImageTk.PhotoImage(image)
        
        img_time = (time.perf_counter() - img_start) * 1000  # TEMP DEBUG
        print(f"[DEBUG] Background image load/resize: {img_time:.2f}ms")  # TEMP DEBUG
        
        # Keep reference to prevent garbage collection
        root.bg_image = bg_image
        
        # Create background label
        background_label = tk.Label(root, image=bg_image)
        background_label.place(relwidth=1, relheight=1)
        print(f"[DEBUG] Background label created")  # TEMP DEBUG
        
        # Set welcome message
        if first_name is None:
            first_name = username
        message = "Welcome to the Makerspace!" if userstatus == 1 else f"Welcome back, {first_name}!"
        
        # --- Layout: position message higher if training data will be shown ---
        if training_status and userstatus == 0:
            message_rely = 0.18
        else:
            message_rely = 0.5
        
        # Create message label
        message_label = tk.Label(root, text=message, font=("Helvetica", 40, "bold"), fg="white", bg="black")
        message_label.place(relx=0.5, rely=message_rely, anchor="center")
        print(f"[DEBUG] Message label created: {message}")  # TEMP DEBUG
        
        # --- Training Status Display ---
        if training_status and userstatus == 0:
            print(f"[DEBUG] Rendering training status display (minimal={MINIMAL_TRAINING_DISPLAY})...")
            
            # Build the list of courses to display based on mode
            required = training_status.get("required", [])
            priority = training_status.get("priority", [])
            optional = training_status.get("optional", [])
            
            # Create a frame for training info
            training_frame = tk.Frame(root, bg="black", bd=0)
            training_frame.place(relx=0.5, rely=0.55, anchor="center")
            
            # --- Required Trainings (always shown) ---
            if required:
                req_label = tk.Label(training_frame, text="Required Trainings",
                                     font=("Helvetica", 20, "bold"), fg="#F56600", bg="black")
                req_label.pack(pady=(5, 2))
                for course in required:
                    icon = "\u2714" if course.get("completed") else "\u2718"
                    color = "#00FF00" if course.get("completed") else "#FF4444"
                    course_label = tk.Label(training_frame,
                                            text=f"  {icon}  {course['name']}",
                                            font=("Helvetica", 16), fg=color, bg="black",
                                            anchor="w")
                    course_label.pack(fill="x", padx=20)
            
            # --- Priority courses (e.g. 3D Printing) shown under Required, no separate heading ---
            if priority:
                for course in priority:
                    icon = "\u2714" if course.get("completed") else "\u2718"
                    color = "#00FF00" if course.get("completed") else "#FF4444"
                    course_label = tk.Label(training_frame,
                                            text=f"  {icon}  {course['name']}",
                                            font=("Helvetica", 16), fg=color, bg="black",
                                            anchor="w")
                    course_label.pack(fill="x", padx=20)
            
            # --- Optional Equipment (only shown when not in minimal mode) ---
            if not MINIMAL_TRAINING_DISPLAY and optional:
                opt_label = tk.Label(training_frame, text="Equipment Trainings",
                                     font=("Helvetica", 20, "bold"), fg="#F56600", bg="black")
                opt_label.pack(pady=(10, 2))
                for course in optional:
                    icon = "\u2714" if course.get("completed") else "\u2718"
                    color = "#00FF00" if course.get("completed") else "#FF4444"
                    course_label = tk.Label(training_frame,
                                            text=f"  {icon}  {course['name']}",
                                            font=("Helvetica", 16), fg=color, bg="black",
                                            anchor="w")
                    course_label.pack(fill="x", padx=20)
            
            # --- Summary line ---
            if MINIMAL_TRAINING_DISPLAY:
                # Count only displayed courses in minimal mode
                displayed = required + priority
                total = len(displayed)
                completed = sum(1 for c in displayed if c.get("completed"))
            else:
                total = training_status.get("total_courses", 0)
                completed = training_status.get("completed_courses", 0)
            
            summary_label = tk.Label(training_frame,
                                     text=f"Completed: {completed}/{total}",
                                     font=("Helvetica", 18, "bold"), fg="white", bg="black")
            summary_label.pack(pady=(12, 2))
            
            # --- Last Updated timestamp ---
            # Prefer fetch_time (actual API call time) over last_db_update (database write time)
            last_updated = training_status.get("fetch_time") or training_status.get("last_db_update")
            if last_updated:
                updated_label = tk.Label(training_frame,
                                         text=f"Last updated: {last_updated}",
                                         font=("Helvetica", 11), fg="#888888", bg="black")
                updated_label.pack(pady=(0, 5))
            
            print(f"[DEBUG] Training status rendered ({completed}/{total} complete)")
        
        # Force updates to display everything
        root.update()
        root.update_idletasks()
        print(f"[DEBUG] Popup displayed successfully")  # TEMP DEBUG
        
        # Function to close popup and process any queued scans
        def close_popup_and_process_queue():
            set_popup_active(False)  # Clear popup flag
            
            # If a scan came in during the popup, process it now
            queued_id = get_queued_scan()
            if queued_id:
                print(f"[QUEUE] Processing queued scan: {queued_id}")
                # Launch new subprocess for queued scan
                subprocess.Popen([sys.executable, __file__, str(queued_id)])
            
            # Destroy the window immediately so it disappears even if
            # background threads (scraping, API calls) are still running
            try:
                root.destroy()
            except:
                pass
        
        # Popup duration: longer when training data is shown to give time to read
        popup_duration = 4000 if (training_status and userstatus == 0) else 2500
        root.after(popup_duration, close_popup_and_process_queue)
        
        popup_total = (time.perf_counter() - popup_create_start) * 1000  # TEMP DEBUG
        print(f"[DEBUG] Total popup creation: {popup_total:.2f}ms")  # TEMP DEBUG
        
    except Exception as e:
        print(f"[ERROR] Exception in show_welcome_popup: {e}")
        import traceback
        traceback.print_exc()
        set_popup_active(False)  # Clear flag on error
        raise

def close_on_escape(event):
    print("Escape key pressed. Exiting program...")
    sys.exit()  # Exit the program

def show_error_popup(message):
    # Create a new window for the error popup
    error_root = tk.Tk()
    error_root.title("Error")
    error_root.attributes('-fullscreen', True)  # Make it fullscreen

    # Set the background color
    error_root.configure(bg="black")

    # Create a label to display the error message
    error_label = tk.Label(error_root, text=message, font=("Helvetica", 40, "bold"), fg="white", bg="black")
    error_label.place(relx=0.5, rely=0.5, anchor="center")  # Center the message

    # Close the popup after 3 seconds
    error_root.after(3000, error_root.destroy)

    # Start the application loop
    error_root.mainloop()

def submit_username(event, entry, root):
    global username
    username = entry.get().strip()
    # Remove @clemson.edu if user typed full email
    if "@clemson.edu" in username:
        username = username.split("@")[0]
    root.quit()

def prompt_for_username():
    global username
    username = None
    
    # Initialize the main window
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")

    root = ctk.CTk()
    root.title("Username Entry")
    
    # Use fullscreen instead of geometry-based sizing.
    # With DPI awareness set at process startup, winfo_screenwidth() returns the
    # true resolution (2560) but CTk's geometry() internally multiplies by the DPI
    # scaling factor, resulting in an oversized window. Fullscreen avoids this.
    root.attributes('-fullscreen', True)
    root.attributes('-topmost', True)
    
    # Force the window to update and then configure it properly
    root.update_idletasks()
    root.lift()
    root.focus_force()
    
    # Create a centered frame for content (don't use pack with expand for main container)
    center_frame = ctk.CTkFrame(master=root, width=800, height=400)
    center_frame.place(relx=0.5, rely=0.5, anchor="center")

    # Title label
    title_label = ctk.CTkLabel(master=center_frame, text="Welcome to the Makerspace!\nEnter Your Clemson Username:", 
                              font=("Arial", 48), text_color="#F56600")
    title_label.pack(pady=30, padx=50)

    # Instruction label
    label = ctk.CTkLabel(master=center_frame, text="The part before the @clemson.edu", 
                        font=("Arial", 32))
    label.pack(pady=10, padx=50)

    # Create an entry box
    entry = ctk.CTkEntry(master=center_frame, width=600, height=60, 
                        placeholder_text="Enter username", font=("Arial", 28))
    entry.pack(pady=20, padx=50)
    
    # Function to check if input is a 6-digit hardware ID
    def check_hardware_id_input(event=None):
        global username, pending_hardware_id
        current_text = entry.get().strip()
        
        # If exactly 6 digits are entered, treat as hardware ID scan
        if current_text.isdigit() and len(current_text) == 6:
            pending_hardware_id = current_text
            username = None
            root.quit()
            return
        
        # Otherwise treat as username submission
        if event and event.keysym == 'Return':
            submit_username(event, entry, root)
    
    # Monitor text changes for hardware ID detection
    entry.bind('<KeyRelease>', check_hardware_id_input)
    
    # Bind escape key to close - still record attendance without username
    def on_escape(event):
        global username
        if not username or username.strip() == "":
            # Escape pressed without username - record scan anyway
            username = ""  # Empty string to indicate no username provided
        root.quit()
    
    root.bind('<Escape>', on_escape)
    
    # Bind the Enter key to submit the form
    root.bind('<Return>', check_hardware_id_input)
    
    # Force focus to the entry box with multiple attempts
    def set_focus():
        root.focus_force()
        root.lift()
        entry.focus_set()
        entry.focus_force()
        root.after(50, lambda: entry.focus_force())
    
    root.after(100, set_focus)
    root.after(300, set_focus)  # Try again after longer delay
    
    # Auto-close after timeout - still record attendance without username
    timeout_id = None
    def timeout_close():
        global username
        if not username or username.strip() == "":
            # Timeout without username - record scan anyway
            username = ""  # Empty string to indicate no username provided
        root.quit()
    
    timeout_id = root.after(25000, timeout_close)
    
    root.mainloop()
    
    # Cancel any pending callbacks before destroying
    if timeout_id:
        try:
            root.after_cancel(timeout_id)
        except:
            pass
    
    root.destroy()
    return username

def scrape_user_thread(username, callback):
    def run():
        first_name, last_name, major = scrape_user(username)
        callback(first_name, last_name, major)
    thread = threading.Thread(target=run)
    thread.start()
    return thread

def fetch_training_status_async(username, callback):
    """Fetch training status in a separate thread to avoid blocking UI"""
    def run():
        if BRIDGE_API_AVAILABLE:
            try:
                fetch_start = time.perf_counter()  # API DEBUG
                status = get_all_training_status(username)
                fetch_time = (time.perf_counter() - fetch_start) * 1000  # API DEBUG
                print(f"[API DEBUG] Async fetch time: {fetch_time:.0f}ms")  # API DEBUG
                callback(status)
            except Exception as e:
                fetch_time = (time.perf_counter() - fetch_start) * 1000  # API DEBUG
                print(f"[API DEBUG] Async fetch time (error): {fetch_time:.0f}ms")  # API DEBUG
                print(f"Error fetching training status: {e}")
                callback(None)
        else:
            callback(None)
    thread = threading.Thread(target=run)
    thread.start()

def main():
    global pending_hardware_id
    pending_hardware_id = None  # Reset at start of each scan
    
    # Clean up any stale flag files on startup
    cleanup_stale_flags()
    
    start_time = time.perf_counter()  # TEMP DEBUG: Start timing
    print(f"\n{'='*60}")  # TEMP DEBUG
    print(f"[DEBUG] SCAN STARTED at {datetime.now().strftime('%H:%M:%S.%f')[:-3]}")  # TEMP DEBUG
    
    hardware_id = sys.argv[1]
    print(f"[DEBUG] Hardware ID: {hardware_id}")  # TEMP DEBUG
    
    # Check if a popup is currently active (across processes)
    if is_popup_active():
        print(f"[QUEUE] Popup is active, queuing scan: {hardware_id}")
        queue_scan(hardware_id)
        print(f"[QUEUE] Scan queued successfully, will process when popup closes")
        return  # Exit this instance, queued scan will be processed when popup closes
    
    # OPTIMIZATION: Try database lookup first (single indexed query)
    db_start = time.perf_counter()  # TEMP DEBUG
    user = None
    try:
        user = db.get_user_by_hardware_id(hardware_id)
        db_time = (time.perf_counter() - db_start) * 1000  # TEMP DEBUG
        print(f"[DEBUG] Database lookup: {db_time:.2f}ms - Found: {user['username'] if user else 'None'}")  # TEMP DEBUG
    except Exception as e:
        db_time = (time.perf_counter() - db_start) * 1000  # TEMP DEBUG
        print(f"[DEBUG] Database lookup failed ({db_time:.2f}ms): {e}")  # TEMP DEBUG
        print(f"Database lookup failed, will use Excel: {e}")
    
    # OPTIMIZATION: Only load Excel if database lookup failed (fallback)
    # This saves ~850ms on every scan when database is working
    workbook, sheet, sheet2 = None, None, None
    if user is None:
        excel_start = time.perf_counter()  # TEMP DEBUG
        workbook, sheet, sheet2 = load_excel()
        excel_time = (time.perf_counter() - excel_start) * 1000  # TEMP DEBUG
        print(f"[DEBUG] Excel load (fallback): {excel_time:.2f}ms")  # TEMP DEBUG
    else:
        print(f"[DEBUG] Excel load: SKIPPED (using database)")  # TEMP DEBUG
    
    # Get username from database or Excel
    username = user['username'] if user else find_hardware_id(sheet2, hardware_id)
    
    ui_start = time.perf_counter()  # TEMP DEBUG
    root = tk.Tk()
    root.withdraw()  # Hide initially
    root.bind("<Escape>", close_on_escape)
    ui_time = (time.perf_counter() - ui_start) * 1000  # TEMP DEBUG
    print(f"[DEBUG] UI init: {ui_time:.2f}ms")  # TEMP DEBUG

    def on_scrape_complete(first_name, last_name, major):
        add_user_to_sheet(sheet_name, sheet2_name, hardware_id, username, first_name, last_name, major, None, 1)
        # Fetch and save training data synchronously (we're already in a background thread)
        if BRIDGE_API_AVAILABLE:
            try:
                print(f"[Background] Fetching training data for new user {username}...")
                status = get_all_training_status(username)
                if status:
                    save_training_data_to_excel(username, status)
                    print(f"[Background] Training data saved for new user {username}")
            except Exception as e:
                print(f"[Background] Error fetching training for new user {username}: {e}")

    if username is not None and username != "":
        # EXISTING USER - Get data from database (fast) or Excel (fallback)
        print(f"User found: {username}")
        
        lookup_start = time.perf_counter()  # TEMP DEBUG
        if user:
            # Fast path: All data from single database query
            first_name = user['first_name']
            last_name = user['last_name']
            major = user['major']
            training_status = db.get_training_data(username)
            if training_status:
                print(f"[API DEBUG] Loaded cached training data from database (no API calls)")
            lookup_time = (time.perf_counter() - lookup_start) * 1000  # TEMP DEBUG
            print(f"[DEBUG] Fast database lookup: {lookup_time:.2f}ms")  # TEMP DEBUG
            print(f"Fast database lookup complete for {username}")
        else:
            # Slow path: Multiple Excel iterations
            first_name, last_name, major = find_userdata(hardware_id, sheet2)
            training_status = get_training_data_from_excel(username, sheet2)
            if training_status:
                print(f"[API DEBUG] Loaded cached training data from Excel (no API calls)")
            lookup_time = (time.perf_counter() - lookup_start) * 1000  # TEMP DEBUG
            print(f"[DEBUG] Excel fallback lookup: {lookup_time:.2f}ms")  # TEMP DEBUG
            print(f"Excel fallback lookup complete for {username}")
        
        save_start = time.perf_counter()  # TEMP DEBUG
        add_user_to_sheet(sheet_name, sheet2_name, hardware_id, username, first_name, last_name, major, None, 0)
        save_time = (time.perf_counter() - save_start) * 1000  # TEMP DEBUG
        print(f"[DEBUG] Save scan record: {save_time:.2f}ms")  # TEMP DEBUG
        
        popup_start = time.perf_counter()  # TEMP DEBUG
        # Show welcome popup immediately with cached data
        root.deiconify()  # Show root window
        make_fullscreen_on_top(root)  # Make fullscreen
        show_welcome_popup(root, username, first_name, 0, training_status)
        popup_time = (time.perf_counter() - popup_start) * 1000  # TEMP DEBUG
        print(f"[DEBUG] Popup display: {popup_time:.2f}ms")  # TEMP DEBUG
        
        # Update training data in background (after display)
        training_thread = update_training_data_async(username)
        
        total_time = (time.perf_counter() - start_time) * 1000  # TEMP DEBUG
        print(f"[DEBUG] TOTAL TIME TO DISPLAY: {total_time:.2f}ms")  # TEMP DEBUG
        print(f"{'='*60}\n")  # TEMP DEBUG
        
        root.mainloop()
        
        # Wait for training data update to finish so the process doesn't exit early
        if training_thread and training_thread.is_alive():
            print("[DEBUG] Waiting for training data update to complete...")
            training_thread.join(timeout=30)
    elif username == "":
        # Username is blank (timeout/escape) - but still record scan for attendance
        print("Username blank - recording scan for attendance without username")
        
        # Check if this hardware ID has scanned before with blank username
        # If so, prompt for username again like a new user
        if user and (not user['username'] or user['username'].strip() == ""):
            print("Hardware ID has previous blank scans - prompting for username again")
            username = None  # Treat as new user
        else:
            # Record scan with blank username for attendance tracking
            now = datetime.now()
            timestamp = now.strftime('%m/%d/%Y %H:%M:%S')
            db.add_scan(hardware_id, "", timestamp, Location)
            
            # Also add to Excel scans sheet
            try:
                with safe_excel_write(file_path) as wb:
                    scans_sheet = wb[sheet_name]
                    scans_sheet.append([int(hardware_id), "", timestamp])
                print(f"Attendance recorded for hardware ID {hardware_id} without username")
            except Exception as e:
                print(f"ERROR saving to Excel: {e}")
                print(f"Attendance recorded in database for hardware ID {hardware_id}")
            
            total_time = (time.perf_counter() - start_time) * 1000  # TEMP DEBUG
            print(f"[DEBUG] TOTAL TIME (blank username): {total_time:.2f}ms")  # TEMP DEBUG
            print(f"{'='*60}\n")  # TEMP DEBUG
            return
    
    # Handle new user or re-prompt for blank username scenarios
    if username is None:
        print("New user detected. Prompting for username.")
        
        # Destroy the original root before prompt_for_username creates its own CTk window.
        # Having two Tk root windows corrupts the event loop, timers, and DPI state.
        try:
            root.destroy()
        except Exception:
            pass
        
        prompt_start = time.perf_counter()  # TEMP DEBUG
        username = prompt_for_username()
        prompt_time = (time.perf_counter() - prompt_start) * 1000  # TEMP DEBUG
        print(f"[DEBUG] Username prompt completed: {prompt_time:.2f}ms")  # TEMP DEBUG
        print(f'Username entered: {username}')
        
        # Check if a hardware ID was scanned during username entry
        if pending_hardware_id is not None:
            print(f"Hardware ID {pending_hardware_id} scanned during username entry. Processing new scan...")
            # Process the new hardware ID scan - use the correct path
            import subprocess
            card_reader_path = get_resource_path("CardReaderMakerspace.py")
            subprocess.Popen([sys.executable, card_reader_path, pending_hardware_id])
            return
        
        if username:
            new_user_start = time.perf_counter()  # TEMP DEBUG
            
            # Create a FRESH Tk root for the popup (same pattern as returning user).
            # The previous root was destroyed before prompt_for_username to avoid
            # dual-Tk corruption. This fresh root has clean event loop and DPI state.
            root = tk.Tk()
            root.withdraw()
            root.bind("<Escape>", close_on_escape)
            root.deiconify()
            make_fullscreen_on_top(root)
            show_welcome_popup(root, username, None, 1, None)
            
            # Scrape user data and fetch training in background (does not touch UI)
            scrape_thread = scrape_user_thread(username, on_scrape_complete)
            
            total_time = (time.perf_counter() - start_time) * 1000  # TEMP DEBUG
            print(f"[DEBUG] TOTAL TIME TO DISPLAY (NEW USER): {total_time:.2f}ms")  # TEMP DEBUG
            print(f"{'='*60}\n")  # TEMP DEBUG
            
            # Process pending events to ensure popup is displayed
            root.update()
            root.mainloop()
            
            # Wait for scrape to finish so the process doesn't exit early
            if scrape_thread and scrape_thread.is_alive():
                print("[DEBUG] Waiting for scrape to complete...")
                scrape_thread.join(timeout=30)
        elif username == "":
            # Timeout/escape with no username - record attendance anyway
            print("Username prompt skipped - recording attendance without username")
            now = datetime.now()
            timestamp = now.strftime('%m/%d/%Y %H:%M:%S')
            db.add_scan(hardware_id, "", timestamp, Location)
            
            try:
                with safe_excel_write(file_path) as wb:
                    scans_sheet = wb[sheet_name]
                    scans_sheet.append([int(hardware_id), "", timestamp])
                print(f"Attendance recorded for hardware ID {hardware_id} without username")
            except Exception as e:
                print(f"ERROR saving to Excel: {e}")
                print(f"Attendance recorded in database for hardware ID {hardware_id}")
            
            total_time = (time.perf_counter() - start_time) * 1000  # TEMP DEBUG
            print(f"[DEBUG] TOTAL TIME (blank username from prompt): {total_time:.2f}ms")  # TEMP DEBUG
            print(f"{'='*60}\n")  # TEMP DEBUG
        else:
            print('Username Prompt cancelled or failed')

if __name__ == "__main__":
    # Initialize database sync system (if enabled)
    if SYNC_AVAILABLE:
        try:
            from database_sync import init_sync
            init_sync()
        except Exception as e:
            print(f"Note: Database sync initialization failed: {e}")
    
    # Perform bidirectional Excel ↔ Database sync on startup
    if BIDIRECTIONAL_SYNC_AVAILABLE:
        try:
            sync_on_startup()
            # Register shutdown sync to run when program exits
            atexit.register(sync_on_shutdown)
        except Exception as e:
            print(f"Note: Bidirectional sync failed: {e}")
    
    main()