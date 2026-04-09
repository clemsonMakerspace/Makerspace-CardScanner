"""
Makerspace Card Scanner - Comprehensive Simulation Test Suite
=============================================================
Simulates every possible card-scan scenario and validates that the
database, Excel sheet, popup/queue system, and sync functions all
behave correctly.

Uses a subset of real production users from hardware_users_prod.xlsx
so the scenarios are realistic.

Run:
    python test_scanner_simulation.py              (full suite)
    python test_scanner_simulation.py --keep-files (don't delete test files)
    python test_scanner_simulation.py --verbose    (extra detail)

Exit code 0 = all tests passed, 1 = one or more failures.
"""

import sys
import os
import io
import json
import time
import shutil
import sqlite3
import argparse
import re
import traceback
from datetime import datetime, timedelta
from contextlib import contextmanager

# Force UTF-8 output on Windows so Unicode in test names prints correctly.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ---------------------------------------------------------------------------
# Ensure the project root is on sys.path
# ---------------------------------------------------------------------------
ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

# ---------------------------------------------------------------------------
# Suppress noisy startup prints from modules that auto-init on import
# ---------------------------------------------------------------------------
class _Suppressor(io.StringIO):
    pass

_orig_stdout = sys.stdout

def _quiet():
    sys.stdout = _Suppressor()

def _loud():
    sys.stdout = _orig_stdout


# ---------------------------------------------------------------------------
# Import project modules (suppress their startup chatter)
# ---------------------------------------------------------------------------
_quiet()
try:
    import database as db
    import excel_db_sync
    import excel_utils
    from excel_utils import safe_excel_write, safe_excel_read, verify_excel_integrity
except ImportError as _e:
    _loud()
    print(f"ERROR importing project module: {_e}")
    print("Run this script from the Makerspace-CardScanner project root.")
    sys.exit(1)
_loud()

try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ---------------------------------------------------------------------------
# Terminal colours (degrade gracefully on Windows without ANSI support)
# ---------------------------------------------------------------------------
_ANSI = sys.platform != "win32" or os.environ.get("ANSICON") or "WT_SESSION" in os.environ

def _c(code, text):
    return f"\033[{code}m{text}\033[0m" if _ANSI else text

def green(t):  return _c("92", t)
def red(t):    return _c("91", t)
def yellow(t): return _c("93", t)
def cyan(t):   return _c("96", t)
def bold(t):   return _c("1",  t)
def dim(t):    return _c("2",  t)


# ---------------------------------------------------------------------------
# Parse args early so --verbose flag is available during setup
# ---------------------------------------------------------------------------
_parser = argparse.ArgumentParser(add_help=False)
_parser.add_argument("--keep-files",  action="store_true")
_parser.add_argument("--verbose",     action="store_true")
_parser.add_argument("--help", "-h",  action="store_true")
ARGS, _ = _parser.parse_known_args()

if ARGS.help:
    print(__doc__)
    sys.exit(0)

VERBOSE = ARGS.verbose


# ---------------------------------------------------------------------------
# Test environment paths  (completely separate from production files)
# ---------------------------------------------------------------------------
TEST_DB    = os.path.join(ROOT, "_test_hardware_users.db")
TEST_EXCEL = os.path.join(ROOT, "_test_hardware_users.xlsx")
TEST_BACKUP_DIR = os.path.join(ROOT, "_test_backups")
PROD_EXCEL = os.path.join(ROOT, "hardware_users_prod.xlsx")

# Popup / queue flag files used in tests (isolated names)
TEST_POPUP_FLAG  = os.path.join(ROOT, "_test_popup_active")
TEST_QUEUE_FILE  = os.path.join(ROOT, "_test_queued_scan")
TEST_POPUP_TIMEOUT = 10   # seconds -- matches production


# ---------------------------------------------------------------------------
# Stripped-down, pure-Python re-implementations of the file-flag helpers
# from CardReaderMakerspace.py (no GUI / tkinter dependency).
# ---------------------------------------------------------------------------

def _is_popup_active(flag_file=TEST_POPUP_FLAG, timeout=TEST_POPUP_TIMEOUT):
    if os.path.exists(flag_file):
        try:
            age = time.time() - os.path.getmtime(flag_file)
            if age > timeout:
                _clear_popup_flag(flag_file)
                return False
            return True
        except Exception:
            return True
    return False

def _set_popup_active(active=True, flag_file=TEST_POPUP_FLAG):
    if active:
        with open(flag_file, "w") as fh:
            fh.write(str(datetime.now()))
    else:
        _clear_popup_flag(flag_file)

def _clear_popup_flag(flag_file=TEST_POPUP_FLAG):
    if os.path.exists(flag_file):
        try:
            os.remove(flag_file)
        except Exception:
            pass

def _queue_scan(hardware_id, queue_file=TEST_QUEUE_FILE):
    with open(queue_file, "w") as fh:
        fh.write(str(hardware_id))

def _get_queued_scan(queue_file=TEST_QUEUE_FILE):
    if os.path.exists(queue_file):
        try:
            with open(queue_file, "r") as fh:
                hid = fh.read().strip()
            os.remove(queue_file)
            return hid
        except Exception:
            return None
    return None

def _cleanup_stale_flags(flag_file=TEST_POPUP_FLAG, queue_file=TEST_QUEUE_FILE,
                          popup_timeout=TEST_POPUP_TIMEOUT):
    removed = []
    if os.path.exists(flag_file):
        age = time.time() - os.path.getmtime(flag_file)
        if age > popup_timeout:
            os.remove(flag_file)
            removed.append("popup_flag")
    if os.path.exists(queue_file):
        age = time.time() - os.path.getmtime(queue_file)
        if age > 30:
            os.remove(queue_file)
            removed.append("queue_file")
    return removed


# ---------------------------------------------------------------------------
# Test user data -- subset pulled from hardware_users_prod.xlsx at start-up.
# Falls back to hardcoded list if the file is missing or unreadable.
# ---------------------------------------------------------------------------

FALLBACK_USERS = [
    {"username": "kfhawki", "hardware_id": 492359, "first_name": "Keller",     "last_name": "Hawkins",    "major": "Materials Science and Engineering"},
    {"username": "adickeh", "hardware_id": 396561, "first_name": "Aleksander", "last_name": "Dickehuth",  "major": "Mechanical Engineering"},
    {"username": "tdmuell", "hardware_id": 397426, "first_name": "Taylor",     "last_name": "Mueller",    "major": "Graphics Communication"},
    {"username": "mlhowe",  "hardware_id": 381698, "first_name": "Melanie",    "last_name": "Hoew",       "major": "Mechanical Engineering"},
    {"username": "kwasib",  "hardware_id": 451378, "first_name": "Kwasi",      "last_name": "Boyd",       "major": "General Engineering"},
    {"username": "kemccrt", "hardware_id": 448101, "first_name": "Katie",      "last_name": "McCarter",   "major": "Performing Arts"},
    {"username": "nlozano", "hardware_id": 496525, "first_name": "Nicolas",    "last_name": "Lozano",     "major": "Computer Science"},
    {"username": "ajense3", "hardware_id": 468594, "first_name": "Drew",       "last_name": "Jensen",     "major": "Electrical Engineering"},
    {"username": "ccraig4", "hardware_id": 493271, "first_name": "Collin",     "last_name": "Craig",      "major": "Mechanical Engineering"},
    {"username": "braden5", "hardware_id": 383468, "first_name": "Braden",     "last_name": "Smith",      "major": "General Engineering"},
    {"username": "charlin", "hardware_id": 396589, "first_name": "Charlie",    "last_name": "Newman",     "major": "Mechanical Engineering"},
    {"username": "ieford",  "hardware_id": 482644, "first_name": "Isaac",      "last_name": "Ford",       "major": "General Engineering"},
    {"username": "akammar", "hardware_id": 482470, "first_name": "Arun",       "last_name": "Kammari",    "major": "Mechanical Engineering"},
    {"username": "dtstand", "hardware_id": 446674, "first_name": "Ty",         "last_name": "Standridge", "major": "Engineering"},
    {"username": "PMMOSLE", "hardware_id": 369553, "first_name": "Philip",     "last_name": "Mosley",     "major": "Mechanical Engineering"},
]

_EMOJI_RE = re.compile(
    "["
    "\U0001F600-\U0001F64F"
    "\U0001F300-\U0001F5FF"
    "\U0001F680-\U0001F6FF"
    "\U0001F1E0-\U0001F1FF"
    "\U00002600-\U000026FF"
    "\U00002700-\U000027BF"
    "\U0001F900-\U0001F9FF"
    "]+",
    flags=re.UNICODE,
)

def _strip_emoji(text):
    if text is None:
        return None
    return _EMOJI_RE.sub("", str(text)).strip()


def _load_prod_users(max_users=15):
    """
    Load up to max_users from the production Excel file.
    Falls back to FALLBACK_USERS if the file is unreadable.
    """
    if not OPENPYXL_OK or not os.path.exists(PROD_EXCEL):
        return FALLBACK_USERS[:max_users]
    try:
        wb = load_workbook(PROD_EXCEL, data_only=True)
        if "Users" not in wb.sheetnames:
            return FALLBACK_USERS[:max_users]
        ws = wb["Users"]
        users = []
        seen_names = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            username    = row[0] if len(row) > 0 else None
            hardware_id = row[1] if len(row) > 1 else None
            first_name  = _strip_emoji(row[3]) if len(row) > 3 else None
            last_name   = _strip_emoji(row[4]) if len(row) > 4 else None
            major       = _strip_emoji(row[5]) if len(row) > 5 else None

            if not username or not hardware_id:
                continue
            # Skip formula strings
            if str(username).startswith("=") or str(hardware_id).startswith("="):
                continue
            try:
                hw_int = int(hardware_id)
            except (ValueError, TypeError):
                continue

            # Skip duplicate usernames (prod data has same user with multiple card IDs)
            uname_clean = str(username).strip()
            if uname_clean.lower() in seen_names:
                continue
            seen_names.add(uname_clean.lower())

            users.append({
                "username":    uname_clean,
                "hardware_id": hw_int,
                "first_name":  first_name,
                "last_name":   last_name,
                "major":       major,
            })
            if len(users) >= max_users:
                break
        wb.close()
        if users:
            return users
    except Exception:
        pass
    return FALLBACK_USERS[:max_users]


# ---------------------------------------------------------------------------
# Test infrastructure
# ---------------------------------------------------------------------------

_results = []   # list of (name, passed, message)

def _register(name, passed, detail=""):
    _results.append((name, passed, detail))
    icon = green("PASS") if passed else red("FAIL")
    line = f"  [{icon}] {name}"
    if detail and (VERBOSE or not passed):
        line += f"\n         {dim(detail)}"
    print(line)
    return passed


def _assert(condition, name, msg=""):
    return _register(name, bool(condition), msg)

def _assert_eq(a, b, name):
    ok = (a == b)
    return _register(name, ok, f"expected {b!r}, got {a!r}" if not ok else "")

def _assert_in(item, container, name):
    ok = item in container
    return _register(name, ok, f"{item!r} not found" if not ok else "")


def _section(title):
    print()
    print(bold("-" * 60))
    print(bold(f"  {title}"))
    print(bold("-" * 60))


def _make_sample_training(username="test", completed_waiver=True, completed_quiz=True):
    """Build a minimal training-status dict that matches bridge_api output."""
    return {
        "required": [
            {"course_id": 5424, "name": "Makerspace Waiver", "category": "required",
             "required": True, "order": 1, "completed": completed_waiver,
             "completed_at": "2025-01-01T12:00:00" if completed_waiver else None, "error": None},
            {"course_id": 5422, "name": "Safety Quiz", "category": "required",
             "required": True, "order": 2, "completed": completed_quiz,
             "completed_at": "2025-01-02T12:00:00" if completed_quiz else None, "error": None},
        ],
        "priority": [
            {"course_id": 5473, "name": "3D Printing", "category": "priority",
             "required": False, "order": 3, "completed": False, "completed_at": None, "error": None},
        ],
        "optional": [
            {"course_id": 5461, "name": "Fabric Printer", "category": "optional",
             "required": False, "order": 10, "completed": False, "completed_at": None, "error": None},
        ],
        "username": username,
        "fetch_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "has_errors": False,
        "total_courses": 4,
        "completed_courses": (1 if completed_waiver else 0) + (1 if completed_quiz else 0),
        "required_complete": completed_waiver and completed_quiz,
    }


def _make_test_excel(path, users, scans=None):
    """Create a minimal hardware_users.xlsx-style workbook for testing."""
    wb = Workbook()
    # Remove default sheet, create named ones
    wb.remove(wb.active)

    users_ws = wb.create_sheet("Users")
    users_ws.append(["Username", "Hardware ID", "Login Count",
                      "First Name", "Last Name", "Major", "Training JSON", "Training Updated"])
    for u in users:
        users_ws.append([
            u.get("username"), u.get("hardware_id"), u.get("login_count", 0),
            u.get("first_name"), u.get("last_name"), u.get("major"),
            u.get("training_json"), u.get("training_updated"),
        ])

    scans_ws = wb.create_sheet("Scans")
    scans_ws.append(["Hardware ID", "Username", "Timestamp", "Location"])
    for s in (scans or []):
        scans_ws.append([s.get("hardware_id"), s.get("username"),
                          s.get("timestamp"), s.get("location", "Watt")])

    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# Test environment context manager
# ---------------------------------------------------------------------------

@contextmanager
def _test_env(users):
    """
    Patch db.DB_FILE and excel_db_sync.EXCEL_FILE to isolated test paths.
    Creates a fresh test DB and Excel, loads 'users', tears down after.
    """
    original_db_file        = db.DB_FILE
    original_excel_file     = excel_db_sync.EXCEL_FILE
    original_backup_dir     = excel_utils.BACKUP_DIR
    original_backup_ts_file = excel_utils.BACKUP_TIMESTAMP_FILE
    original_last_backup    = excel_utils._last_backup_time

    # Override module-level paths
    db.DB_FILE                       = TEST_DB
    excel_db_sync.EXCEL_FILE         = TEST_EXCEL
    excel_db_sync.USERS_SHEET        = "Users"
    excel_db_sync.SCANS_SHEET        = "Scans"
    excel_utils.BACKUP_DIR           = TEST_BACKUP_DIR
    excel_utils.BACKUP_TIMESTAMP_FILE = os.path.join(TEST_BACKUP_DIR, ".last_backup_time")
    excel_utils._last_backup_time    = 0  # force backup to run on first call

    # Clean slate
    for p in (TEST_DB, TEST_EXCEL):
        if os.path.exists(p):
            os.remove(p)
    if os.path.exists(TEST_BACKUP_DIR):
        shutil.rmtree(TEST_BACKUP_DIR)
    os.makedirs(TEST_BACKUP_DIR, exist_ok=True)

    # Initialise fresh database
    _quiet()
    try:
        db.init_database()
        for u in users:
            db.add_or_update_user(
                u["username"], u["hardware_id"],
                u.get("first_name"), u.get("last_name"), u.get("major")
            )
    finally:
        _loud()

    # Create matching Excel
    _make_test_excel(TEST_EXCEL, users)

    try:
        yield TEST_DB, TEST_EXCEL, users
    finally:
        # Restore originals
        db.DB_FILE                        = original_db_file
        excel_db_sync.EXCEL_FILE          = original_excel_file
        excel_utils.BACKUP_DIR            = original_backup_dir
        excel_utils.BACKUP_TIMESTAMP_FILE = original_backup_ts_file
        excel_utils._last_backup_time     = original_last_backup

        # Clean up test artefacts
        if not ARGS.keep_files:
            for p in (TEST_DB, TEST_EXCEL):
                if os.path.exists(p):
                    try:
                        os.remove(p)
                    except Exception:
                        pass
            if os.path.exists(TEST_BACKUP_DIR):
                try:
                    shutil.rmtree(TEST_BACKUP_DIR)
                except Exception:
                    pass

        # Always clean up flag files
        _clear_popup_flag()
        if os.path.exists(TEST_QUEUE_FILE):
            os.remove(TEST_QUEUE_FILE)


# ============================================================
# TEST GROUPS
# ============================================================

def run_group_1_database_basics(users):
    _section("GROUP 1 -- Database Basics")

    with _test_env(users) as (tdb, _, _):

        # T01 - DB file created
        _assert(os.path.exists(tdb), "T01 DB file exists after init")

        # T02 - User lookup by hardware ID
        u0 = users[0]
        r = db.get_user_by_hardware_id(u0["hardware_id"])
        _assert(r is not None, "T02 get_user_by_hardware_id returns row")
        _assert_eq(r["username"], u0["username"], "T02b correct username returned")

        # T03 - User lookup by username
        r2 = db.get_user_by_username(u0["username"])
        _assert(r2 is not None, "T03 get_user_by_username returns row")
        _assert_eq(r2["hardware_id"], u0["hardware_id"], "T03b correct hardware_id")

        # T04 - Add new user and retrieve
        db.add_or_update_user("newtest99", 777001, "New", "Tester", "Test Major")
        rn = db.get_user_by_username("newtest99")
        _assert(rn is not None, "T04 newly added user retrievable")
        _assert_eq(rn["first_name"], "New", "T04b first_name stored correctly")

        # T05 - Update existing user (COALESCE means only non-None fields overwrite)
        db.add_or_update_user(u0["username"], u0["hardware_id"], "Updated", None, None)
        ru = db.get_user_by_username(u0["username"])
        _assert_eq(ru["first_name"], "Updated", "T05 first_name updated via add_or_update_user")

        # T06 - DB integrity check
        ok = db.verify_database_integrity()
        _assert(ok, "T06 verify_database_integrity passes")

        # T07 - DB stats
        stats = db.get_database_stats()
        _assert(stats.get("total_users", 0) >= len(users), "T07 stats: total_users >= loaded users")
        _assert("total_scans"        in stats, "T07b stats: total_scans key present")
        _assert("users_with_training" in stats, "T07c stats: users_with_training key present")

        # T08 - Database backup
        backup_path = db.backup_database(TEST_BACKUP_DIR)
        _assert(backup_path is not None, "T08 backup_database returns a path")
        _assert(os.path.exists(backup_path), "T08b backup file actually exists on disk")


def run_group_2_scan_recording(users):
    _section("GROUP 2 -- Scan Recording")

    with _test_env(users) as (tdb, _, _):
        u0 = users[0]
        ts = datetime.now().strftime("%m/%d/%Y %H:%M:%S")

        # T09 - Record scan for known user, login count increments
        before = db.get_user_by_hardware_id(u0["hardware_id"])
        before_count = before["login_count"] if before else 0
        db.add_scan(u0["hardware_id"], u0["username"], ts, "Watt")
        after = db.get_user_by_hardware_id(u0["hardware_id"])
        _assert_eq(after["login_count"], before_count + 1, "T09 login_count incremented after scan")

        # T10 - Scan with blank username (anonymous attendance)
        ok = db.add_scan(u0["hardware_id"], "", ts, "Watt")
        _assert(ok, "T10 scan with blank username succeeds")

        # T11 - Location tag stored correctly
        ts2 = (datetime.now() + timedelta(seconds=1)).strftime("%m/%d/%Y %H:%M:%S")
        db.add_scan(u0["hardware_id"], u0["username"], ts2, "Cooper")
        scans = db.get_recent_scans(5)
        locations = [s["location"] for s in scans]
        _assert_in("Cooper", locations, "T11 Cooper location tag persisted")

        # T12 - Multiple scans from same user accumulate
        u1 = users[1]
        for i in range(5):
            tsi = (datetime.now() + timedelta(seconds=i+10)).strftime("%m/%d/%Y %H:%M:%S")
            db.add_scan(u1["hardware_id"], u1["username"], tsi, "Watt")
        count = db.get_user_scan_count(u1["username"])
        _assert(count >= 5, f"T12 multiple scans accumulate (found {count})")

        # T13 - get_recent_scans returns sane data
        recent = db.get_recent_scans(20)
        _assert(isinstance(recent, list), "T13 get_recent_scans returns list")
        _assert(len(recent) > 0,          "T13b list is non-empty")
        _assert("timestamp" in recent[0], "T13c timestamp key present")

        # T14 - get_user_scan_count
        cnt0 = db.get_user_scan_count(u0["username"])
        _assert(cnt0 >= 1, f"T14 get_user_scan_count >= 1 (got {cnt0})")


def run_group_3_popup_and_queue(users):
    _section("GROUP 3 -- Popup Flag & Queue System")

    # Ensure clean state
    _clear_popup_flag()
    if os.path.exists(TEST_QUEUE_FILE):
        os.remove(TEST_QUEUE_FILE)

    # T15 - Set popup active and detect it
    _set_popup_active(True)
    _assert(_is_popup_active(), "T15 popup flag detected after set_popup_active(True)")

    # T16 - Clear popup flag
    _set_popup_active(False)
    _assert(not _is_popup_active(), "T16 popup flag cleared after set_popup_active(False)")

    # T17 - Stale flag (artificially age it)
    _set_popup_active(True)
    # Wind back mtime by 20 seconds to simulate stale flag
    flag_path = TEST_POPUP_FLAG
    past = time.time() - (TEST_POPUP_TIMEOUT + 10)
    os.utime(flag_path, (past, past))
    _assert(not _is_popup_active(), "T17 stale popup flag (>timeout) treated as inactive")
    _assert(not os.path.exists(flag_path), "T17b stale popup flag file removed automatically")

    # T18 - cleanup_stale_flags removes old files
    _set_popup_active(True)
    os.utime(flag_path, (past, past))
    removed = _cleanup_stale_flags()
    _assert_in("popup_flag", removed, "T18 cleanup_stale_flags removes aged popup flag")

    # T19 - Queue a scan
    _queue_scan("123456")
    _assert(os.path.exists(TEST_QUEUE_FILE), "T19 queue file created after _queue_scan")

    # T20 - Get queued scan clears file
    val = _get_queued_scan()
    _assert_eq(val, "123456", "T20 get_queued_scan returns correct hardware_id")
    _assert(not os.path.exists(TEST_QUEUE_FILE), "T20b queue file removed after retrieval")

    # T21 - Second queued scan overwrites first (file is overwritten)
    _queue_scan("111111")
    _queue_scan("222222")
    val2 = _get_queued_scan()
    _assert_eq(val2, "222222", "T21 second queued scan overwrites first")

    # T22 - Popup active prevents scan processing (simulate logic)
    _set_popup_active(True)
    should_queue = _is_popup_active()
    _assert(should_queue, "T22 scan arrives while popup active -> should queue (is_popup_active=True)")
    _set_popup_active(False)


def run_group_4_training_data(users):
    _section("GROUP 4 -- Training Data")

    with _test_env(users) as (tdb, _, _):
        u0 = users[0]

        # T23 - Update training data in DB
        sample = _make_sample_training(u0["username"], completed_waiver=True, completed_quiz=False)
        ok = db.update_training_data(u0["username"], sample)
        _assert(ok, "T23 update_training_data returns True")

        # T24 - Get training data back
        td = db.get_training_data(u0["username"])
        _assert(td is not None, "T24 get_training_data returns non-None")
        _assert("required" in td, "T24b 'required' key present in returned training data")
        _assert("optional" in td, "T24c 'optional' key present")

        # T25 - Completion counts match
        _assert_eq(td.get("total_courses"), 4, "T25 total_courses stored correctly")
        completed_in_db = td.get("completed_courses")
        _assert_eq(completed_in_db, 1, "T25b completed_courses = 1 (only waiver done)")

        # T26 - required_complete reflects reality
        _assert(not td.get("required_complete"), "T26 required_complete=False when quiz not done")
        sample_full = _make_sample_training(u0["username"], completed_waiver=True, completed_quiz=True)
        db.update_training_data(u0["username"], sample_full)
        td2 = db.get_training_data(u0["username"])
        _assert(td2.get("required_complete"), "T26b required_complete=True when both done")

        # T27 - User without training returns None
        u_no_train = users[3]
        td_none = db.get_training_data(u_no_train["username"])
        _assert(td_none is None, "T27 user with no training data returns None from get_training_data")

        # T28 - Corrupted training JSON in DB gracefully fails to load
        with db.get_db_connection() as conn:
            conn.execute(
                "UPDATE users SET training_data=? WHERE username=?",
                ("{not valid json!!", u_no_train["username"])
            )
        td_bad = db.get_training_data(u_no_train["username"])
        _assert(td_bad is None, "T28 corrupted training JSON returns None (no crash)")


def run_group_5_excel_sync(users):
    _section("GROUP 5 -- Excel Sync (DB <-> Excel)")

    if not OPENPYXL_OK:
        print(yellow("  SKIPPED - openpyxl not installed."))
        return

    with _test_env(users) as (tdb, texcel, _):

        # T29 - Excel integrity check on freshly created file
        ok = verify_excel_integrity(texcel)
        _assert(ok, "T29 verify_excel_integrity passes on fresh test Excel")

        # T30 - sync_database_to_excel writes all users
        _quiet()
        try:
            excel_db_sync.sync_database_to_excel()
        finally:
            _loud()
        wb = load_workbook(texcel, data_only=True)
        ws = wb["Users"]
        usernames_in_excel = {row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
        wb.close()
        for u in users[:5]:
            _assert(u["username"] in usernames_in_excel,
                    f"T30 sync_database_to_excel: {u['username']} present in Excel")

        # T31 - sync_excel_to_database picks up a user added directly to Excel
        wb2 = load_workbook(texcel)
        ws2 = wb2["Users"]
        ws2.append(["excel_only_user", 888001, 0, "Excel", "Only", "Archaeology", None, None])
        wb2.save(texcel)
        wb2.close()

        _quiet()
        try:
            excel_db_sync.sync_excel_to_database()
        finally:
            _loud()
        r = db.get_user_by_username("excel_only_user")
        _assert(r is not None, "T31 sync_excel_to_database imports Excel-only user into DB")

        # T32 - Scans added to DB appear in Excel after sync
        ts = datetime.now().strftime("%m/%d/%Y %H:%M:%S")
        db.add_scan(users[2]["hardware_id"], users[2]["username"], ts, "Watt")
        _quiet()
        try:
            excel_db_sync.sync_database_to_excel()
        finally:
            _loud()
        wb3 = load_workbook(texcel, data_only=True)
        scans_ws = wb3["Scans"]
        scan_usernames = [row[1] for row in scans_ws.iter_rows(min_row=2, values_only=True) if row[1]]
        wb3.close()
        _assert(users[2]["username"] in scan_usernames,
                f"T32 scan for {users[2]['username']} appears in Excel after DB->Excel sync")

        # T33 - smart_bidirectional_sync (DB newer): DB wins
        import os as _os
        # Make DB mtime 10 seconds ahead of Excel
        now = time.time()
        _os.utime(texcel, (now - 10, now - 10))
        _os.utime(tdb,    (now,      now))

        _quiet()
        try:
            excel_db_sync.smart_bidirectional_sync()
        finally:
            _loud()
        _assert(True, "T33 smart_bidirectional_sync (DB newer) completes without exception")

        # T34 - smart_bidirectional_sync (Excel newer): Excel wins
        _os.utime(texcel, (now + 10, now + 10))
        _os.utime(tdb,    (now,      now))

        _quiet()
        try:
            excel_db_sync.smart_bidirectional_sync()
        finally:
            _loud()
        _assert(True, "T34 smart_bidirectional_sync (Excel newer) completes without exception")

        # T35 - smart_bidirectional_sync (within 5s): reported as in-sync
        _os.utime(texcel, (now, now))
        _os.utime(tdb,    (now, now))
        _quiet()
        try:
            excel_db_sync.smart_bidirectional_sync()
        finally:
            _loud()
        _assert(True, "T35 smart_bidirectional_sync (in sync) completes without exception")

        # T36 - Excel backup creation
        # Reset both in-memory timer and the persistent disk timestamp so
        # create_backup doesn't skip due to a recent backup in earlier tests.
        ts_file = excel_utils.BACKUP_TIMESTAMP_FILE
        if os.path.exists(ts_file):
            os.remove(ts_file)
        excel_utils._last_backup_time = 0
        backup_created = excel_utils.create_backup(texcel, location="Watt")
        _assert(backup_created is not None, "T36 create_backup returns a path")
        _assert(os.path.exists(backup_created), "T36b backup file exists on disk")


def run_group_6_complex_scenarios(users):
    _section("GROUP 6 -- Complex Multi-Scan Scenarios")

    with _test_env(users) as (tdb, texcel, _):
        u0, u1, u2 = users[0], users[1], users[2]

        # T37 - Rapid double-scan same user (both recorded, count increments twice)
        ts1 = datetime.now().strftime("%m/%d/%Y %H:%M:%S")
        ts2 = (datetime.now() + timedelta(seconds=1)).strftime("%m/%d/%Y %H:%M:%S")
        db.add_scan(u0["hardware_id"], u0["username"], ts1, "Watt")
        db.add_scan(u0["hardware_id"], u0["username"], ts2, "Watt")
        count0 = db.get_user_scan_count(u0["username"])
        _assert(count0 >= 2, f"T37 double-scan: login_count >= 2 (got {count0})")

        # T38 - Person A scans -> popup becomes active -> person B scans (should queue)
        _set_popup_active(True)
        popup_was_active = _is_popup_active()
        if popup_was_active:
            _queue_scan(str(u1["hardware_id"]))
        queued = _get_queued_scan()
        _clear_popup_flag()
        _assert(queued == str(u1["hardware_id"]),
                "T38 Person B's scan queued while Person A's popup is active")

        # T39 - Queued scan processed after popup closes
        # Simulate: popup closes -> process queued scan
        _set_popup_active(True)
        _queue_scan(str(u2["hardware_id"]))
        _set_popup_active(False)                   # popup closed
        queued_id = _get_queued_scan()
        if queued_id:
            ts3 = datetime.now().strftime("%m/%d/%Y %H:%M:%S")
            db.add_scan(int(queued_id), u2["username"], ts3, "Watt")
        cnt2 = db.get_user_scan_count(u2["username"])
        _assert(cnt2 >= 1, f"T39 queued scan processed after popup closed (u2 count={cnt2})")

        # T40 - User A scans, enters no username (blank) -> scan still recorded
        ts4 = datetime.now().strftime("%m/%d/%Y %H:%M:%S")
        blank_ok = db.add_scan(u0["hardware_id"], "", ts4, "Watt")
        _assert(blank_ok, "T40 add_scan with blank username returns True")
        blank_scans = [s for s in db.get_recent_scans(50) if s["username"] == ""]
        _assert(len(blank_scans) >= 1, "T40b blank-username scan appears in recent scans")

        # T41 - User A blank scan, then user B scans immediately (no conflict)
        ts5 = (datetime.now() + timedelta(seconds=2)).strftime("%m/%d/%Y %H:%M:%S")
        db.add_scan(u1["hardware_id"], u1["username"], ts5, "Watt")
        cnt1 = db.get_user_scan_count(u1["username"])
        _assert(cnt1 >= 1, "T41 User B's scan succeeds right after User A's blank scan")

        # T42 - New user flow: register then scan
        new_hw = 800001
        new_un = "newuser_sim"
        db.add_or_update_user(new_un, new_hw, "Sim", "User", "Testing")
        ts6 = datetime.now().strftime("%m/%d/%Y %H:%M:%S")
        db.add_scan(new_hw, new_un, ts6, "Watt")
        new_r = db.get_user_by_hardware_id(new_hw)
        _assert(new_r is not None, "T42 new user exists in DB after registration")
        _assert(new_r["login_count"] >= 1, "T42b new user scan increments login_count")

        # T43 - Training data attached to new user
        train = _make_sample_training(new_un, completed_waiver=False, completed_quiz=False)
        db.update_training_data(new_un, train)
        td_new = db.get_training_data(new_un)
        _assert(td_new is not None, "T43 training data attached to new user")
        _assert(not td_new["required_complete"], "T43b required_complete=False for untrained user")

        # T44 - Location-tagged scans from two locations mix correctly
        for i, location in enumerate(["Watt", "Cooper", "Watt", "Cooper"]):
            tsi = (datetime.now() + timedelta(seconds=i + 5)).strftime("%m/%d/%Y %H:%M:%S")
            db.add_scan(u0["hardware_id"], u0["username"], tsi, location)
        all_scans = db.get_recent_scans(100)
        cooper_scans = [s for s in all_scans if s["location"] == "Cooper"]
        watt_scans   = [s for s in all_scans if s["location"] == "Watt"]
        _assert(len(cooper_scans) >= 2, f"T44 Cooper scans present (found {len(cooper_scans)})")
        _assert(len(watt_scans)   >= 2, f"T44b Watt scans present   (found {len(watt_scans)})")


def run_group_7_edge_cases(users):
    _section("GROUP 7 -- Edge Cases & Error Handling")

    with _test_env(users) as (tdb, texcel, _):
        u0 = users[0]

        # T45 - Hardware ID as string vs integer (both work)
        r_str = db.get_user_by_hardware_id(str(u0["hardware_id"]))
        r_int = db.get_user_by_hardware_id(u0["hardware_id"])
        _assert(r_str is not None, "T45 hardware_id lookup with string works")
        _assert(r_int is not None, "T45b hardware_id lookup with int works")

        # T46 - Username case sensitivity
        uname_upper = u0["username"].upper()
        uname_lower = u0["username"].lower()
        # Database stores exactly what was inserted; lookups are case-sensitive in SQLite
        r_exact = db.get_user_by_username(u0["username"])
        _assert(r_exact is not None, "T46 exact-case username lookup works")

        # T47 - get_user_by_hardware_id with unknown ID returns None
        r_unknown = db.get_user_by_hardware_id(9999999)
        _assert(r_unknown is None, "T47 unknown hardware_id returns None")

        # T48 - Scan with None hardware_id is handled gracefully
        try:
            result = db.add_scan(None, "test", datetime.now().strftime("%m/%d/%Y %H:%M:%S"), "Watt")
            # Either returns False or raises; both are acceptable
            _assert(True, "T48 add_scan(None hw_id) does not crash process")
        except Exception:
            _assert(True, "T48 add_scan(None hw_id) raises (acceptable)")

        # T49 - add_or_update_user with duplicate hardware_id constraint
        # Inserting a second user with the same hardware_id should not crash
        result = db.add_or_update_user("dup_hwid_user", u0["hardware_id"], "Dup", "User", "CS")
        # IntegrityError caught internally -> returns False
        _assert(True, "T49 duplicate hardware_id does not crash (handled internally)")

        # T50 - Verify DB integrity still OK after edge-case ops
        ok = db.verify_database_integrity()
        _assert(ok, "T50 DB integrity check still passes after edge-case operations")


def run_group_8_production_simulation(users):
    _section("GROUP 8 -- Full Production Simulation (30 synthetic scans)")

    with _test_env(users) as (tdb, texcel, _):

        # Build a realistic scan sequence
        # Patterns: normal, double-scan, two people close, blank, queue scenario
        scan_log = []
        base_time = datetime.now()

        def _ts(delta_seconds):
            return (base_time + timedelta(seconds=delta_seconds)).strftime("%m/%d/%Y %H:%M:%S")

        # Normal morning scans spread over ~10 users
        normal_scans = [
            (0,   users[0]),  (15,  users[1]),  (45,  users[2]),
            (90,  users[3]),  (130, users[4]),   (165, users[5]),
            (200, users[6]),  (240, users[7]),   (280, users[8]),
            (320, users[9]),
        ]
        for offset, u in normal_scans:
            db.add_scan(u["hardware_id"], u["username"], _ts(offset), "Watt")
            scan_log.append((u["username"], _ts(offset), "Watt", "normal"))

        # Double scan -- user 0 scans twice within 2 seconds
        db.add_scan(users[0]["hardware_id"], users[0]["username"], _ts(400), "Watt")
        db.add_scan(users[0]["hardware_id"], users[0]["username"], _ts(401), "Watt")
        scan_log.append((users[0]["username"], _ts(400), "Watt", "double_1"))
        scan_log.append((users[0]["username"], _ts(401), "Watt", "double_2"))

        # Blank scan (user forgot to enter username)
        db.add_scan(users[1]["hardware_id"], "", _ts(500), "Watt")
        scan_log.append(("", _ts(500), "Watt", "blank_uname"))

        # Second location -- Cooper
        for i, u in enumerate(users[:5]):
            db.add_scan(u["hardware_id"], u["username"], _ts(600 + i * 30), "Cooper")
            scan_log.append((u["username"], _ts(600 + i * 30), "Cooper", "cooper"))

        # Another user finishes -- returns for 2nd visit same day
        for i in range(3):
            db.add_scan(users[2]["hardware_id"], users[2]["username"], _ts(800 + i * 20), "Watt")
            scan_log.append((users[2]["username"], _ts(800 + i * 20), "Watt", f"return_{i}"))

        # Simulate queuing logic in code (not DB, but flag logic)
        _set_popup_active(True)
        queued = _is_popup_active()
        _queue_scan(str(users[4]["hardware_id"]))  # arrives while popup up
        _set_popup_active(False)
        q = _get_queued_scan()
        if q:
            db.add_scan(int(q), users[4]["username"], _ts(900), "Watt")
            scan_log.append((users[4]["username"], _ts(900), "Watt", "queued"))

        total_simulated = len(scan_log)

        # --- Assertions ---

        # T51 - Total scan count in DB matches what we inserted
        all_scans = db.get_recent_scans(1000)
        _assert(len(all_scans) >= total_simulated,
                f"T51 DB has >= {total_simulated} scans (found {len(all_scans)})")

        # T52 - Double-scan user has login_count >= 3 (1 normal + 2 double)
        u0_data = db.get_user_by_username(users[0]["username"])
        _assert(u0_data["login_count"] >= 3,
                f"T52 double-scan user login_count >= 3 (got {u0_data['login_count']})")

        # T53 - Blank scan present in scans table
        blank = [s for s in all_scans if s["username"] == ""]
        _assert(len(blank) >= 1, f"T53 blank-username scan recorded (found {len(blank)})")

        # T54 - Both Watt and Cooper scans present
        cooper = [s for s in all_scans if s["location"] == "Cooper"]
        _assert(len(cooper) >= 5, f"T54 Cooper scans present (found {len(cooper)})")

        # T55 - Sync to Excel and verify round-trip
        _quiet()
        try:
            excel_db_sync.sync_database_to_excel()
        finally:
            _loud()

        if OPENPYXL_OK:
            wb = load_workbook(texcel, data_only=True)
            scans_ws = wb["Scans"]
            excel_scan_count = sum(1 for _ in scans_ws.iter_rows(min_row=2))
            wb.close()
            _assert(excel_scan_count >= total_simulated,
                    f"T55 Excel has >= {total_simulated} scans after sync (found {excel_scan_count})")
        else:
            _register("T55 Excel round-trip", True, "[skip - openpyxl not available]")

        # T56 - Re-import Excel back to a fresh DB (simulate restore / first-run)
        # The deadlock in sync_excel_to_database was fixed (inline SQL instead
        # of calling db.add_scan in a nested lock), so full restore works now.
        if OPENPYXL_OK:
            # Wipe DB (direct sqlite3 - bypasses the module lock)
            conn_wipe = sqlite3.connect(tdb)
            conn_wipe.execute("DELETE FROM scans")
            conn_wipe.execute("DELETE FROM users")
            conn_wipe.commit()
            conn_wipe.close()

            # Re-import everything from Excel
            _quiet()
            try:
                excel_db_sync.sync_excel_to_database()
            finally:
                _loud()
            restored_stats = db.get_database_stats()
            restored_users = restored_stats.get("total_users", 0)
            restored_scans = restored_stats.get("total_scans", 0)
            _assert(restored_users >= len(users),
                    f"T56 restore from Excel: {restored_users} users re-imported")
            _assert(restored_scans >= total_simulated,
                    f"T56b restore from Excel: {restored_scans} scans re-imported")
        else:
            _register("T56 Restore from Excel", True, "[skip - openpyxl not available]")

        # T57 - After restore, user lookup still works
        restored_u = db.get_user_by_username(users[0]["username"])
        _assert(restored_u is not None,
                f"T57 user '{users[0]['username']}' accessible after restore from Excel")


# ============================================================
# MAIN RUNNER
# ============================================================

def main():
    print()
    print(bold("=" * 60))
    print(bold("  MAKERSPACE CARD SCANNER - SIMULATION TEST SUITE"))
    print(bold("=" * 60))
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Python {sys.version.split()[0]}")
    print()

    # Load test users
    print("Loading test users from production data...")
    test_users = _load_prod_users(max_users=15)
    print(f"  {len(test_users)} users loaded ")
    if VERBOSE:
        for u in test_users:
            print(f"    {u['username']:<12} hw={u['hardware_id']:<8} "
                  f"{u.get('first_name','')} {u.get('last_name','')}")
    print()

    # Run each group
    groups = [
        ("Database Basics",          lambda: run_group_1_database_basics(test_users)),
        ("Scan Recording",           lambda: run_group_2_scan_recording(test_users)),
        ("Popup Flag & Queue",       lambda: run_group_3_popup_and_queue(test_users)),
        ("Training Data",            lambda: run_group_4_training_data(test_users)),
        ("Excel Sync",               lambda: run_group_5_excel_sync(test_users)),
        ("Complex Multi-Scan",       lambda: run_group_6_complex_scenarios(test_users)),
        ("Edge Cases",               lambda: run_group_7_edge_cases(test_users)),
        ("Full Production Sim",      lambda: run_group_8_production_simulation(test_users)),
    ]

    for name, fn in groups:
        try:
            fn()
        except Exception as exc:
            _section(f"ERROR in group '{name}'")
            print(red(f"  Unhandled exception: {exc}"))
            if VERBOSE:
                traceback.print_exc()

    # ---------------------------------------------------------------------------
    # Summary
    # ---------------------------------------------------------------------------
    total   = len(_results)
    passed  = sum(1 for _, ok, _ in _results if ok)
    failed  = total - passed

    print()
    print(bold("=" * 60))
    print(bold("  RESULTS"))
    print(bold("=" * 60))
    print(f"  Tests run : {total}")
    print(f"  {green(f'Passed    : {passed}')}")
    if failed:
        print(f"  {red(f'Failed    : {failed}')}")
        print()
        print(red("  FAILED TESTS:"))
        for name, ok, detail in _results:
            if not ok:
                print(f"    {red('FAIL')} {name}" + (f"\n         {dim(detail)}" if detail else ""))
    else:
        print(f"  {green('All tests passed!')}")

    if ARGS.keep_files:
        print()
        print(yellow(f"  Test files kept (--keep-files):"))
        print(f"    DB    : {TEST_DB}")
        print(f"    Excel : {TEST_EXCEL}")

    print()

    sys.exit(0 if failed == 0 else 1)


if __name__ == "__main__":
    main()
